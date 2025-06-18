"""
作者：Mison
邮箱：1360962086@qq.com
仓库：https://github.com/MisonL/dify_chat_tester
许可证：MIT

Dify 聊天客户端测试工具

如何获取 Dify API 密钥和应用 ID：
1. 登录 Dify 控制台：https://cloud.dify.ai/
2. 选择或创建应用
3. 进入应用设置页面
4. 在 "API 密钥" 部分查看或创建 API 密钥
5. 应用 ID 可以在应用概览页面的 URL 中找到，格式为：
   https://cloud.dify.ai/app/{应用ID}/configuration
   或直接在应用设置页面查看

使用步骤：
1. 安装依赖：
   确保已安装以下 Python 库。如果未安装，请在终端运行：
   pip install requests openpyxl
2. 运行程序：
   python dify_chat_tester.py
3. 按照提示输入 Dify 基础 URL、API 密钥和应用 ID
4. 选择角色
5. 选择运行模式：
   - '1' 进入会话模式（实时对话）
   - '2' 进入批量询问模式（通过 Excel 文件批量处理）
6. 会话模式命令：
   - 'exit' 退出程序
   - '/new' 开启新的对话（重置上下文）
7. 批量询问模式：
   - **支持从 Dify 应用 URL 中自动提取应用 ID 和基础 URL。**
   - 按照提示选择或输入 Excel 文件路径（支持当前目录下的文件或完整路径）。
   - 按照提示输入问题所在列的名称或序号。
   - 按照提示输入回答结果保存列的名称或序号（**支持新增列**）。
   - 按照提示选择是否在控制台显示每个问题的回答内容。
   - 按照提示输入每个请求之间的间隔时间（秒）。
   - 处理完成后，显示详细的统计信息。
8. 查看生成的 Excel 文件：
   - 会话模式：`chat_log.xlsx` (记录每次对话的详细信息)。
   - 批量询问模式：
     - `batch_query_log_YYYYMMDD_HHMMSS.xlsx` (此文件是**详细操作日志**，记录每次请求的时间戳、原始问题、Dify响应、成功状态、错误信息和对话ID)。
     - **原始 Excel 文件** (Dify的回答结果会**直接写入**您选择的原始文件中的指定列，并实时保存)。

注意事项：
- 确保 Dify API 支持流式响应。
- 角色列表可在代码中 `ROLES` 变量中扩展。
- 程序会自动创建 Excel 文件记录对话。
- 私有化部署用户可输入自定义 API 地址，程序会尝试从 URL 中自动提取应用 ID。
- Dify API 密钥必须以 `app-` 开头，程序会进行校验。
- 支持多轮对话（通过 `conversation_id` 维护上下文）。
- 支持随时开启新对话（输入 `/new` 重置上下文）。
- 批量询问模式下，请确保 Excel 文件格式正确，问题列名准确，且问题内容不为空。
- 程序在请求 Dify API 时会显示等待指示器，并在收到第一个字符或遇到错误时停止。
- 批量询问模式下，程序会安全地写入 Excel 单元格，处理合并单元格的情况。
- 批量询问模式下，程序会检查问题是否为空，并跳过空问题。
- 批量询问模式下，每处理一个问题后会实时保存原始 Excel 文件，以防数据丢失。
"""

import requests
import json
import openpyxl
from openpyxl.cell.cell import MergedCell # 导入 MergedCell 类型
from datetime import datetime
import os
import sys
import threading
import time

# Excel 配置
CHAT_LOG_FILE_NAME = "chat_log.xlsx"
ROLES = ["员工", "门店"]  # 可扩展更多角色

def write_cell_safely(worksheet, row, col, value):
    """
    安全地写入 Excel 单元格，处理合并单元格的情况。
    如果目标单元格是合并单元格的一部分，则写入合并区域的左上角单元格。
    """
    cell_obj = worksheet.cell(row=row, column=col)
    if isinstance(cell_obj, MergedCell):
        # 如果是合并单元格的一部分，找到其合并区域的左上角单元格
        for merged_range in worksheet.merged_cells.ranges:
            if cell_obj.coordinate in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                worksheet.cell(row=min_row, column=min_col).value = value
                return
    else:
        cell_obj.value = value

def init_excel_log(file_name, headers):
    """初始化 Excel 日志文件"""
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        if worksheet is None:
            worksheet = workbook.create_sheet("Log")
        
        # 设置表头
        worksheet.append(headers)
    
    return workbook, worksheet

def log_to_excel(worksheet, row_data):
    """记录到 Excel"""
    worksheet.append(row_data)

def show_waiting_indicator(stop_event):
    """显示等待状态指示器"""
    indicators = ["⣾", "⣽", "⣻", "⢿", "⡿", "⣟", "⣯", "⣷"]
    idx = 0
    while not stop_event.is_set():
        sys.stdout.write(f"\rDify: 正在思考 {indicators[idx]} ")
        sys.stdout.flush()
        idx = (idx + 1) % len(indicators)
        time.sleep(0.1)
    # 清除等待指示器
    sys.stdout.write("\r" + " " * 30 + "\r")
    sys.stdout.flush()

def send_to_dify(base_url, api_key, app_id, role, user_input, conversation_id=None, stream=True, show_indicator=True):
    """
    发送消息到 Dify API
    支持多轮对话（通过 conversation_id 维护上下文）
    """
    url = f"{base_url}/v1/chat-messages"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "inputs": {
            "role": role  # 传递选择的角色
        },
        "query": user_input,
        "response_mode": "streaming" if stream else "blocking",
        "user": "chat_tester"
    }
    
    # 添加对话ID用于多轮对话
    if conversation_id:
        payload["conversation_id"] = conversation_id
    
    # 初始化变量确保在所有路径中都有定义
    stop_event = threading.Event()
    waiting_thread = None
    first_char_printed = False  # 跟踪是否已打印第一个字符
    new_conversation_id = None  # 存储新对话ID
    
    try:
        if show_indicator:
            # 启动等待状态指示器
            waiting_thread = threading.Thread(target=show_waiting_indicator, args=(stop_event,))
            waiting_thread.daemon = True
            waiting_thread.start()
        
        response = requests.post(url, headers=headers, json=payload, stream=stream)
        response.raise_for_status()
        
        if stream:
            # 处理流式响应
            full_response = ""
            for line in response.iter_lines():
                if line:
                    decoded_line = line.decode('utf-8')
                    if decoded_line.startswith("data:"):
                        try:
                            data = json.loads(decoded_line[5:])
                            
                            # 捕获对话ID（用于多轮对话）
                            if "conversation_id" in data:
                                new_conversation_id = data["conversation_id"]
                            
                            if "answer" in data:
                                # 如果是第一个字符，停止等待指示器
                                if show_indicator and not first_char_printed:
                                    stop_event.set()
                                    if waiting_thread is not None:
                                        waiting_thread.join(timeout=0.5)
                                    sys.stdout.write("Dify: ")
                                    sys.stdout.flush()
                                    first_char_printed = True
                                
                                # 打印响应内容
                                if show_indicator:
                                    print(data["answer"], end="", flush=True)
                                full_response += data["answer"]
                            elif "error" in data:
                                # 遇到错误时停止等待指示器
                                if show_indicator:
                                    stop_event.set()
                                    if waiting_thread is not None:
                                        waiting_thread.join(timeout=0.5)
                                error_msg = data.get("error", "未知错误")
                                print(f"\n错误: {error_msg}", file=sys.stderr)
                                return "", False, error_msg, None
                        except json.JSONDecodeError:
                            # JSON解析错误时停止等待指示器
                            if show_indicator:
                                stop_event.set()
                                if waiting_thread is not None:
                                    waiting_thread.join(timeout=0.5)
                            print(f"\nJSON解析错误: {decoded_line}", file=sys.stderr)
                            return "", False, f"JSON解析错误: {decoded_line}", None
            if show_indicator:
                print()  # 换行
            return full_response, True, None, new_conversation_id
        else:
            # 处理阻塞响应
            data = response.json()
            # 停止等待指示器
            if show_indicator:
                stop_event.set()
                if waiting_thread is not None:
                    waiting_thread.join(timeout=0.5)
            
            # 捕获对话ID（用于多轮对话）
            if "conversation_id" in data:
                new_conversation_id = data["conversation_id"]
            
            if "answer" in data:
                if show_indicator:
                    print("Dify:", data["answer"])
                return data["answer"], True, None, new_conversation_id
            elif "error" in data:
                error_msg = data.get("error", "未知错误")
                if show_indicator:
                    print(f"错误: {error_msg}", file=sys.stderr)
                return "", False, error_msg, None
            else:
                return "", False, "未知响应格式", None
    except requests.exceptions.HTTPError as e:
        # 停止等待指示器
        if show_indicator:
            stop_event.set()
            if waiting_thread is not None:
                waiting_thread.join(timeout=0.5)
        
        error_msg = f"HTTP错误: {e.response.status_code} - {e.response.text}"
        if show_indicator:
            print(f"\r错误: {error_msg}", file=sys.stderr)
        return "", False, error_msg, None
    except Exception as e:
        # 停止等待指示器
        if show_indicator:
            stop_event.set()
            if waiting_thread is not None:
                waiting_thread.join(timeout=0.5)
        
        error_msg = f"请求错误: {str(e)}"
        if show_indicator:
            print(f"\r错误: {error_msg}", file=sys.stderr)
        return "", False, error_msg, None
    finally:
        # 确保在任何情况下都停止等待指示器
        if show_indicator:
            stop_event.set()
            if waiting_thread is not None and waiting_thread.is_alive():
                waiting_thread.join(timeout=0.5)

def run_interactive_chat(base_url, api_key, app_id, selected_role):
    """运行会话模式"""
    # 初始化 Excel
    chat_headers = ["时间戳", "角色", "用户输入", "Dify响应", "是否成功", "错误信息", "对话轮次", "对话ID"]
    workbook, worksheet = init_excel_log(CHAT_LOG_FILE_NAME, chat_headers)
    
    print(f"\n已选择角色: {selected_role}")
    print("命令说明:")
    print("  - 输入 'exit' 退出程序")
    print("  - 输入 '/new' 开启新的对话（重置上下文）\n")
    
    # 多轮对话支持
    conversation_id = None  # 对话ID，用于维护多轮对话上下文
    conversation_round = 0  # 对话轮次计数器
    
    # 聊天循环
    while True:
        user_input = input("你: ")
        
        # 处理退出命令
        if user_input.lower() == "exit":
            break
            
        # 处理开启新对话命令
        if user_input.strip() == "/new":
            conversation_id = None
            conversation_round = 0
            print("\n已开启新对话（上下文已重置）\n")
            continue
        
        conversation_round += 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 发送到 Dify
        response, success, error, new_conversation_id = send_to_dify(
            base_url,
            api_key, # 传递 api_key
            app_id,  # 传递 app_id
            selected_role,
            user_input,
            conversation_id,
            show_indicator=True # 会话模式显示指示器
        )
        
        # 更新对话ID（用于后续多轮对话）
        if new_conversation_id:
            conversation_id = new_conversation_id
        
        # 记录到 Excel
        log_to_excel(
            worksheet, 
            [
                timestamp, 
                selected_role, 
                user_input, 
                response, 
                success, 
                error,
                conversation_round,
                conversation_id or ""  # 确保传递字符串（None时用空字符串）
            ]
        )
    
    # 保存并关闭 Excel
    workbook.save(CHAT_LOG_FILE_NAME)
    print(f"\n对话已保存到 {CHAT_LOG_FILE_NAME} (共 {conversation_round} 轮对话)")

def run_batch_query(base_url, api_key, app_id, selected_role):
    """运行批量询问模式"""
    print("\n--- 进入批量询问模式 ---")
    
    # 列出当前目录下的 Excel 文件
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and os.path.isfile(f)]
    
    selected_excel_file = None
    while True:
        if excel_files:
            print("\n当前目录下的 Excel 文件:")
            for i, file_name in enumerate(excel_files):
                print(f"{i+1}. {file_name}")
            
            file_input = input("请输入 Excel 文件序号或直接输入文件路径: ")
            
            try:
                file_index = int(file_input)
                if 1 <= file_index <= len(excel_files):
                    excel_file_path = excel_files[file_index - 1]
                else:
                    print(f"错误: 无效的文件序号 '{file_index}'。请重新输入。", file=sys.stderr)
                    continue
            except ValueError:
                # 用户输入的是路径
                excel_file_path = file_input
        else:
            excel_file_path = input("当前目录下没有找到 Excel 文件。请输入包含询问内容的 Excel 文件路径: ")

        if not os.path.exists(excel_file_path):
            print(f"错误: 文件 '{excel_file_path}' 不存在。请重新输入。", file=sys.stderr)
            continue
        
        try:
            batch_workbook = openpyxl.load_workbook(excel_file_path)
            batch_worksheet = batch_workbook.active
            if batch_worksheet is None: # 确保工作表不为None
                print(f"错误: Excel 文件 '{excel_file_path}' 中没有活动工作表。请重新输入。", file=sys.stderr)
                continue
            selected_excel_file = excel_file_path
            break # 成功读取文件并获取工作表，跳出循环
        except Exception as e:
            print(f"错误: 无法读取 Excel 文件 '{excel_file_path}'。请确保文件格式正确且未被占用。错误信息: {e}。请重新输入。", file=sys.stderr)
            continue

    # 获取列名
    column_names = [cell.value for cell in batch_worksheet[1]]
    print(f"\n已选择文件: {selected_excel_file}")
    print("\nExcel 文件中的列名:")
    for i, col_name in enumerate(column_names):
        print(f"{i+1}. {col_name}")

    question_col_input = input("请输入问题所在列的名称或序号 (例如: '问题' 或 '2'): ")
    question_col_index = -1
    try:
        # 尝试将输入作为序号处理
        col_num = int(question_col_input)
        if 1 <= col_num <= len(column_names):
            question_col_index = col_num - 1
        else:
            print(f"错误: 无效的列序号 '{col_num}'。请检查输入。", file=sys.stderr)
            return
    except ValueError:
        # 如果不是数字，则作为列名处理
        try:
            question_col_index = column_names.index(question_col_input)
        except ValueError:
            print(f"错误: 未找到列名为 '{question_col_input}' 的列。请检查输入。", file=sys.stderr)
            return

    # --- 新增：选择回答结果保存列 ---
    print("\n请选择回答结果保存列:")
    print("现有列名:")
    for i, col_name in enumerate(column_names):
        print(f"{i+1}. {col_name}")
    answer_col_input = input("请输入要保存回答结果的列名或序号 (例如: '回答' 或直接输入新列名，默认: '结果'): ") or "结果"

    answer_col_index = -1
    try:
        # 尝试将输入作为序号处理
        col_num = int(answer_col_input)
        if 1 <= col_num <= len(column_names):
            answer_col_index = col_num - 1
        else:
            # 如果是无效序号，但用户可能想新增一个名为数字的列
            # 此时将数字视为新的列名
            new_col_name = answer_col_input
            column_names.append(new_col_name)
            write_cell_safely(batch_worksheet, 1, len(column_names), new_col_name)
            answer_col_index = len(column_names) - 1
            print(f"已新增列: '{new_col_name}'")
    except ValueError:
        # 如果不是数字，则作为列名处理
        if answer_col_input in column_names:
            answer_col_index = column_names.index(answer_col_input)
        else:
            # 新增列
            column_names.append(answer_col_input)
            write_cell_safely(batch_worksheet, 1, len(column_names), answer_col_input)
            answer_col_index = len(column_names) - 1
            print(f"已新增列: '{answer_col_input}'")
    # --- 新增结束 ---

    # --- 新增：询问是否显示每个问题的回答内容 ---
    display_response_choice = input("是否在控制台显示每个问题的回答内容？ (y/N，默认: N): ").lower()
    show_batch_response = (display_response_choice == 'y')
    # --- 新增结束 ---

    try:
        request_interval = float(input("请输入每个请求之间的间隔时间（秒，建议 0.1 或更高）: "))
        if request_interval < 0:
            raise ValueError
    except ValueError:
        print("无效的间隔时间，请输入一个非负数字。", file=sys.stderr)
        return

    # 详细日志文件，用于记录每次请求的详细信息
    output_file_name = f"batch_query_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    batch_log_headers = ["时间戳", "角色", "原始问题", "Dify响应", "是否成功", "错误信息", "对话ID"]
    output_workbook, output_worksheet = init_excel_log(output_file_name, batch_log_headers)

    total_queries = 0
    successful_queries = 0
    failed_queries = 0
    start_time = time.time()

    print("\n开始批量询问...")
    for row_idx in range(2, batch_worksheet.max_row + 1): # 从第二行开始读取数据
        question_cell_value = batch_worksheet.cell(row=row_idx, column=question_col_index + 1).value
        question = str(question_cell_value) if question_cell_value is not None else "" # 确保转换为字符串
        
        if not question.strip(): # 检查问题是否为空或只包含空格
            print(f"警告: 第 {row_idx} 行问题为空，跳过。", file=sys.stderr)
            failed_queries += 1 # 空问题也算作失败
            log_to_excel(output_worksheet, [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                selected_role,
                question, # 原始问题为空
                "",
                False,
                "问题为空",
                0,
                ""
            ])
            # 在原始文件中也标记为空
            write_cell_safely(batch_worksheet, row_idx, answer_col_index + 1, "[问题为空]")
            continue # 跳过当前循环的剩余部分

        total_queries += 1 # 只有非空问题才计入总数
        print(f"\n处理问题 (第 {total_queries} 个): {question[:50]}...") # 打印问题，并换行
        
        response, success, error, conversation_id = send_to_dify(
            base_url,
            api_key, # 传递 api_key
            app_id,  # 传递 app_id
            selected_role,
            question,
            stream=True, # 批量模式下启用流式输出
            show_indicator=True # 批量模式下显示指示器
        )
        
        if success:
            successful_queries += 1
            print(f"问题 (第 {total_queries} 个) 处理完成。") # 简洁提示
            # 将Dify响应写入原始Excel文件的指定列
            write_cell_safely(batch_worksheet, row_idx, answer_col_index + 1, response)
            # 流式输出已在 send_to_dify 内部处理，这里不再重复打印
        else:
            failed_queries += 1
            print(f"问题 (第 {total_queries} 个) 处理失败。错误: {error}") # 简洁提示
            # 将错误信息写入原始Excel文件的指定列
            write_cell_safely(batch_worksheet, row_idx, answer_col_index + 1, f"[错误]: {error}")
            # 错误信息已在 send_to_dify 内部打印，这里不再重复打印
        
        # 记录详细日志到新的Excel文件
        log_to_excel(output_worksheet, [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            selected_role,
            question,
            response,
            success,
            error,
            1, # 批量询问通常是单轮对话，这里设为1
            conversation_id or ""
        ])
        
        # 每完成一个问题后保存原始Excel文件，以免丢失
        try:
            batch_workbook.save(excel_file_path)
        except Exception as e:
            print(f"警告: 无法保存原始 Excel 文件 '{excel_file_path}'。请确保文件未被占用。错误信息: {e}", file=sys.stderr)

        time.sleep(request_interval) # 间隔时间

    end_time = time.time()
    total_duration = end_time - start_time
    average_duration = total_duration / total_queries if total_queries > 0 else 0

    # 保存详细日志文件
    output_workbook.save(output_file_name)
    print(f"详细日志已保存到 {output_file_name}")
    print("\n--- 批量询问统计 ---")
    print(f"总处理数量: {total_queries}")
    print(f"成功数量: {successful_queries}")
    print(f"失败数量: {failed_queries}")
    print(f"总用时长: {total_duration:.2f} 秒")
    print(f"平均用时长: {average_duration:.2f} 秒/问题")
    print("--------------------")

import urllib.parse # 导入 URL 解析模块
import re # 导入正则表达式模块

def main():
    """主函数"""
    # 获取用户输入
    print("Dify 配置 (直接回车使用默认值)")
    
    # Dify 基础 URL 和应用 ID 校验与提取
    base_url = ""
    DIFY_APP_ID = ""
    while True:
        dify_url_input = input("请输入 Dify 基础 URL 或应用 URL (例如: https://api.dify.ai 或 https://cloud.dify.ai/app/your-app-id/workflow): ") or "https://api.dify.ai"
        parsed_url = urllib.parse.urlparse(dify_url_input)
        
        # 检查是否是有效的 URL 格式
        if not parsed_url.scheme:
            # 如果没有协议，尝试添加 https://
            dify_url_input = "https://" + dify_url_input
            parsed_url = urllib.parse.urlparse(dify_url_input)

        if not parsed_url.netloc:
            print("错误: Dify URL 格式无效。请确保输入有效的域名或 URL。", file=sys.stderr)
            continue
        
        # 提取基础 URL (scheme + netloc)
        base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"

        # 尝试从路径中提取应用 ID
        app_id_match = re.search(r'/app/([0-9a-fA-F-]{36})', parsed_url.path)
        if app_id_match:
            DIFY_APP_ID = app_id_match.group(1)
            print(f"已从 URL 中提取 Dify 基础 URL 为: {base_url}")
            print(f"已从 URL 中提取 Dify 应用 ID 为: {DIFY_APP_ID}")
            break
        else:
            # 如果没有从 URL 中提取到应用 ID，则单独询问
            print(f"已解析 Dify 基础 URL 为: {base_url}")
            while True:
                DIFY_APP_ID = input("请输入 Dify 应用 ID: ")
                if not DIFY_APP_ID:
                    print("错误: Dify 应用 ID 不能为空。", file=sys.stderr)
                    continue
                # 简单校验应用ID是否为UUID格式（可选，但有助于减少错误）
                if not re.match(r'^[0-9a-fA-F-]{36}$', DIFY_APP_ID):
                    print("警告: Dify 应用 ID 格式可能不正确（非标准 UUID 格式）。请确认。", file=sys.stderr)
                break
            break # 成功获取基础URL和应用ID，跳出循环

    # Dify API 密钥校验
    while True:
        DIFY_API_KEY = input("请输入 Dify API 密钥: ")
        if not DIFY_API_KEY:
            print("错误: Dify API 密钥不能为空。", file=sys.stderr)
            continue
        if not DIFY_API_KEY.startswith("app-"):
            print("错误: Dify API 密钥必须以 'app-' 开头。", file=sys.stderr)
            continue
        break
    
    # 角色选择
    print("\n可用角色:")
    for i, role in enumerate(ROLES, 1):
        print(f"{i}. {role}")
    
    role_choice = int(input("请选择角色 (输入序号): "))
    if role_choice < 1 or role_choice > len(ROLES):
        print("无效的角色选择!", file=sys.stderr)
        return
    selected_role = ROLES[role_choice - 1]
    
    # 模式选择
    print("\n请选择运行模式:")
    print("1. 会话模式 (实时对话)")
    print("2. 批量询问模式 (通过 Excel 文件批量询问)")
    mode_choice = input("请输入模式序号 (1 或 2): ")

    if mode_choice == '1':
        run_interactive_chat(base_url, DIFY_API_KEY, DIFY_APP_ID, selected_role)
    elif mode_choice == '2':
        run_batch_query(base_url, DIFY_API_KEY, DIFY_APP_ID, selected_role)
    else:
        print("无效的模式选择，程序退出。", file=sys.stderr)

if __name__ == "__main__":
    main()
