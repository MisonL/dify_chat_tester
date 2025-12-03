"""Excel Utils 模块的单元测试。"""

import os
import tempfile
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from dify_chat_tester.excel_utils import (
    clean_excel_text,
    init_excel_log,
    log_to_excel,
    write_cell_safely,
)


def test_clean_excel_text():
    """测试清理 Excel 非法字符"""
    # 正常文本
    assert clean_excel_text("Hello World") == "Hello World"
    
    # 包含非法字符
    text_with_illegal = "Hello\x00\x01World\x7f"
    cleaned = clean_excel_text(text_with_illegal)
    assert cleaned == "HelloWorld"
    
    # None 值
    assert clean_excel_text(None) == ""
    
    # 数字
    assert clean_excel_text(123) == "123"


def test_init_excel_log_new_file():
    """测试初始化新的 Excel 日志文件"""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = os.path.join(tmpdir, "test.xlsx")
        headers = ["A", "B", "C"]
        
        workbook, worksheet = init_excel_log(file_path, headers)
        
        assert workbook is not None
        assert worksheet is not None
        # 检查表头
        assert worksheet.cell(1, 1).value == "A"
        assert worksheet.cell(1, 2).value == "B"
        assert worksheet.cell(1, 3).value == "C"
        
        workbook.close()


def test_init_excel_log_existing_file():
    """测试打开已存在的 Excel 文件"""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_path = os.path.join(tmpdir, "test.xlsx")
        
        # 先创建文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Old", "Header"])
        wb.save(file_path)
        wb.close()
        
        # 重新打开
        workbook, worksheet = init_excel_log(file_path, ["New", "Header"])
        
        # 应该保留旧内容
        assert worksheet.cell(1, 1).value == "Old"
        
        workbook.close()


def test_log_to_excel():
    """测试记录到 Excel"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    row_data = ["Data1\x00", "Data2", 123]
    log_to_excel(worksheet, row_data)
    
    # 检查清理后的数据
    assert worksheet.cell(1, 1).value == "Data1"
    assert worksheet.cell(1, 2).value == "Data2"
    assert worksheet.cell(1, 3).value == "123"
    
    workbook.close()


def test_write_cell_safely_normal():
    """测试安全写入普通单元格"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    write_cell_safely(worksheet, 1, 1, "Test")
    
    assert worksheet.cell(1, 1).value == "Test"
    
    workbook.close()


def test_write_cell_safely_merged():
    """测试安全写入合并单元格"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # 合并单元格
    worksheet.merge_cells("A1:B2")
    
    # 写入合并单元格的右下角
    write_cell_safely(worksheet, 2, 2, "Merged Value")
    
    # 值应该在左上角
    assert worksheet.cell(1, 1).value == "Merged Value"
    
    workbook.close()
