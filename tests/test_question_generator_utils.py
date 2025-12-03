"""Question Generator 辅助函数的单元测试"""

import os
import tempfile
from unittest.mock import MagicMock, patch


from dify_chat_tester.question_generator import (
    export_questions_to_excel,
    parse_questions_from_response,
    generate_questions_for_document,
    read_markdown_files,
)


def test_read_markdown_files():
    """测试读取 Markdown 文件"""
    with tempfile.TemporaryDirectory() as tmpdir:
        # 创建测试文件
        file1 = os.path.join(tmpdir, "doc1.md")
        file2 = os.path.join(tmpdir, "doc2.txt")  # 应该被忽略

        with open(file1, "w") as f:
            f.write("# Doc 1 Content")
        with open(file2, "w") as f:
            f.write("Ignored")

        file_names, contents = read_markdown_files(tmpdir)

        assert "doc1.md" in file_names
        assert "doc2.txt" not in file_names
        assert contents["doc1.md"] == "# Doc 1 Content"


def test_read_markdown_files_empty_dir():
    """测试空目录"""
    with tempfile.TemporaryDirectory() as tmpdir:
        file_names, contents = read_markdown_files(tmpdir)
        assert file_names == []
        assert contents == {}


def test_read_markdown_files_not_exist():
    """测试不存在的目录"""
    file_names, contents = read_markdown_files("/non/existent/path")
    assert file_names == []
    assert contents == {}


def test_parse_questions_from_response():
    """测试从文本提取问题"""
    text = """
    Here are some questions:
    1. What is AI?
    2. How does it work?
    - Is it safe?
    * Can it think?
    
    Some other text.
    """
    questions = parse_questions_from_response(text)

    assert "What is AI?" in questions
    assert "How does it work?" in questions
    assert "Is it safe?" in questions
    assert "Can it think?" in questions
    assert len(questions) == 4


def test_parse_questions_from_response_json():
    """测试从 JSON 格式文本提取问题"""
    text = """
    ```json
    [
        "Question 1",
        "Question 2"
    ]
    ```
    """
    questions = parse_questions_from_response(text)
    assert "Question 1" in questions
    assert "Question 2" in questions


def test_generate_questions_for_document():
    """测试为文档生成问题"""
    mock_provider = MagicMock()
    mock_provider.send_message.return_value = ("1. Q1\n2. Q2", True, None, None)

    questions = generate_questions_for_document(
        mock_provider, "model", "role", "doc.md", "content"
    )

    assert "Q1" in questions
    assert "Q2" in questions
    mock_provider.send_message.assert_called_once()


def test_generate_questions_for_document_failure():
    """测试生成失败"""
    mock_provider = MagicMock()
    mock_provider.send_message.return_value = ("", False, "Error", None)

    questions = generate_questions_for_document(
        mock_provider, "model", "role", "doc.md", "content"
    )

    assert questions == []


def test_export_questions_to_excel():
    """测试导出到 Excel"""
    questions = [("doc1.md", "Q1"), ("doc1.md", "Q2")]

    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = os.path.join(tmpdir, "output.xlsx")

        success = export_questions_to_excel(questions, output_file)

        assert success is True
        assert os.path.exists(output_file)

        # 验证内容
        import openpyxl

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "doc1.md"
        assert ws.cell(row=2, column=2).value == "Q1"


def test_export_questions_to_excel_permission_error():
    """测试导出权限错误"""
    # 模拟保存时抛出 PermissionError
    with patch("openpyxl.Workbook.save", side_effect=PermissionError):
        success = export_questions_to_excel([("d", "q")], "output.xlsx")
        assert success is False
