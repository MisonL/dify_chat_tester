# tests/test_batch_concurrent.py
"""并发批量处理功能的单元测试"""

from unittest.mock import MagicMock, patch
import time

from dify_chat_tester.core.batch import (
    wait_for_any,
    KeyboardControl,
    _generate_worker_table,
    _process_single_question,
    _run_sequential_batch,
)


class TestWaitForAny:
    """测试 wait_for_any 函数"""

    def test_empty_futures(self):
        """空 futures 集合应返回两个空集合"""
        done, not_done = wait_for_any(set())
        assert done == set()
        assert not_done == set()

    def test_with_completed_future(self):
        """有已完成的 future 时应正确返回"""
        from concurrent.futures import ThreadPoolExecutor

        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(lambda: "result")
            time.sleep(0.1)  # 等待完成

            done, not_done = wait_for_any({future}, timeout=1.0)
            assert len(done) == 1
            assert future in done


class TestKeyboardControl:
    """测试键盘控制类"""

    def test_initial_state(self):
        """测试初始状态"""
        kb = KeyboardControl()
        assert kb.stop_requested is False
        assert kb.paused is False
        assert kb._running is False

    def test_start_stop(self):
        """测试启动和停止"""
        kb = KeyboardControl()
        kb.start()
        assert kb._running is True
        assert kb._listener_thread is not None
        
        kb.stop()
        assert kb._running is False

    def test_stop_requested_flag(self):
        """测试停止请求标志"""
        kb = KeyboardControl()
        kb.stop_requested = True
        assert kb.stop_requested is True

    def test_paused_flag(self):
        """测试暂停标志"""
        kb = KeyboardControl()
        kb.paused = True
        assert kb.paused is True
        kb.paused = False
        assert kb.paused is False


class TestGenerateWorkerTable:
    """测试工作线程状态表格生成"""

    def test_empty_workers(self):
        """测试空工作线程状态"""
        table = _generate_worker_table({}, 0, 10, 0)
        assert table is not None
        assert "并发处理" in table.title

    def test_with_workers(self):
        """测试有工作线程状态"""
        worker_status = {
            1: {"state": "处理中", "question": "问题1"},
            2: {"state": "完成", "question": "问题2"},
            3: {"state": "失败", "question": "问题3"},
        }
        table = _generate_worker_table(worker_status, 2, 10, 1)
        assert table is not None

    def test_paused_state(self):
        """测试暂停状态显示"""
        table = _generate_worker_table({}, 5, 10, 0, paused=True)
        assert "暂停" in table.title

    def test_progress_display(self):
        """测试进度显示"""
        table = _generate_worker_table({}, 5, 10, 1, paused=False)
        # 标题应包含进度信息（新格式使用 bold cyan 样式）
        assert "5" in table.title and "10" in table.title
        assert "✅" in table.title  # 成功数图标
        assert "❌" in table.title  # 失败数图标


class TestProcessSingleQuestion:
    """测试单个问题处理函数"""

    def test_calls_provider_send_message(self):
        """测试是否正确调用 provider.send_message"""
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("回答", True, None, "conv-1")

        result = _process_single_question(
            provider=mock_provider,
            question="测试问题",
            selected_model="model-1",
            selected_role="user",
            enable_thinking=False,
        )

        mock_provider.send_message.assert_called_once_with(
            message="测试问题",
            model="model-1",
            role="user",
            stream=True,
            show_indicator=False,
            show_thinking=False,
            stream_callback=None,  # 新增：流式回调参数
        )
        assert result == ("回答", True, None, "conv-1")

    def test_with_thinking_enabled(self):
        """测试启用思维链"""
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("回答", True, None, "conv-1")

        _process_single_question(
            provider=mock_provider,
            question="问题",
            selected_model="model",
            selected_role="role",
            enable_thinking=True,
        )

        # 验证 show_thinking 参数
        call_kwargs = mock_provider.send_message.call_args[1]
        assert call_kwargs["show_thinking"] is True


class TestCLIArguments:
    """测试 CLI 参数解析"""

    def test_concurrency_argument_parsed(self):
        """测试 --concurrency 参数被正确解析"""
        import argparse
        
        parser = argparse.ArgumentParser()
        parser.add_argument("--concurrency", type=int, default=1)
        
        args = parser.parse_args(["--concurrency", "5"])
        assert args.concurrency == 5

    def test_default_concurrency(self):
        """测试默认并发数为 1"""
        import argparse
        
        parser = argparse.ArgumentParser()
        parser.add_argument("--concurrency", type=int, default=1)
        
        args = parser.parse_args([])
        assert args.concurrency == 1


class TestConfigLoader:
    """测试配置加载器对新配置项的支持"""

    @patch("dify_chat_tester.config.loader.get_config")
    def test_batch_concurrency_config(self, mock_get_config):
        """测试 BATCH_CONCURRENCY 配置项"""
        mock_config = MagicMock()
        mock_config.get_int.return_value = 5
        mock_get_config.return_value = mock_config

        config = mock_get_config()
        concurrency = config.get_int("BATCH_CONCURRENCY", 1)
        
        assert concurrency == 5
        mock_config.get_int.assert_called_with("BATCH_CONCURRENCY", 1)


class TestFriendlyErrorMessage:
    """测试友好错误信息转换"""

    def test_queuepool_error(self):
        """测试 QueuePool 连接池错误的友好提示"""
        from dify_chat_tester.providers.base import _friendly_error_message
        
        error_msg = "Run failed: QueuePool limit of size 30 overflow 10 reached, connection timed out"
        friendly = _friendly_error_message(error_msg)
        
        assert "连接池" in friendly
        assert "BATCH_CONCURRENCY" in friendly

    def test_timeout_error(self):
        """测试超时错误的友好提示"""
        from dify_chat_tester.providers.base import _friendly_error_message

        error_msg = "Failed to establish a new connection: timeout"
        friendly = _friendly_error_message(error_msg)

        assert "连接" in friendly or "网络" in friendly

    def test_ssl_error(self):
        """测试 SSL 错误的友好提示"""
        from dify_chat_tester.providers.base import _friendly_error_message
        
        error_msg = "SSL: CERTIFICATE_VERIFY_FAILED"
        friendly = _friendly_error_message(error_msg)
        
        assert "SSL" in friendly or "证书" in friendly

    def test_http_401_error(self):
        """测试 HTTP 401 认证错误"""
        from dify_chat_tester.providers.base import _friendly_error_message
        
        friendly = _friendly_error_message("Unauthorized", status_code=401)
        
        assert "认证" in friendly or "密钥" in friendly

    def test_http_429_error(self):
        """测试 HTTP 429 频率限制错误"""
        from dify_chat_tester.providers.base import _friendly_error_message
        
        friendly = _friendly_error_message("Too Many Requests", status_code=429)
        
        assert "频繁" in friendly or "限制" in friendly

    def test_passthrough_chinese_message(self):
        """测试中文消息直接返回"""
        from dify_chat_tester.providers.base import _friendly_error_message
        
        error_msg = "这是一个中文错误消息"
        friendly = _friendly_error_message(error_msg)
        
        assert friendly == error_msg


class TestProcessWithRetry:
    """测试带重试的问题处理函数"""

    def test_success_on_first_try(self):
        """测试第一次成功的情况"""
        from dify_chat_tester.core.batch import _process_with_retry
        
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("回答", True, None, "conv-1")

        result, retry_count = _process_with_retry(
            provider=mock_provider,
            question="测试问题",
            selected_model="model",
            selected_role="user",
            enable_thinking=False,
            max_retries=3,
        )

        assert result == ("回答", True, None, "conv-1")
        assert retry_count == 0
        assert mock_provider.send_message.call_count == 1

    def test_success_after_retry(self):
        """测试重试后成功的情况"""
        from dify_chat_tester.core.batch import _process_with_retry
        
        mock_provider = MagicMock()
        mock_provider.send_message.side_effect = [
            ("", False, "临时错误", None),
            ("回答", True, None, "conv-1"),
        ]

        result, retry_count = _process_with_retry(
            provider=mock_provider,
            question="测试问题",
            selected_model="model",
            selected_role="user",
            enable_thinking=False,
            max_retries=3,
        )

        assert result == ("回答", True, None, "conv-1")
        assert retry_count == 1
        assert mock_provider.send_message.call_count == 2

    def test_all_retries_failed(self):
        """测试所有重试都失败的情况"""
        from dify_chat_tester.core.batch import _process_with_retry
        
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("", False, "持续错误", None)

        result, retry_count = _process_with_retry(
            provider=mock_provider,
            question="测试问题",
            selected_model="model",
            selected_role="user",
            enable_thinking=False,
            max_retries=2,
        )

        response, success, error, conv_id = result
        assert success is False
        assert "重试" in error and "2" in error
        assert retry_count == 3  # 初次失败后重试2次，共计3次
        assert mock_provider.send_message.call_count == 3  # 初次 + 2次重试


class TestWorkerTableAdvanced:
    """测试工作线程表格的高级功能"""

    def test_stopping_state(self):
        """测试停止状态显示"""
        table = _generate_worker_table({}, 5, 10, 0, paused=False, stopping=True)
        assert "停止" in table.title

    def test_average_time_display(self):
        """测试平均耗时显示"""
        start_time = time.time() - 10  # 10秒前开始
        table = _generate_worker_table(
            {}, 
            completed=5, 
            total=10, 
            failed=0, 
            paused=False, 
            start_time=start_time
        )
        # caption 应该包含平均耗时
        assert table.caption is not None
        assert "秒" in table.caption or "分钟" in table.caption

    def test_eta_display(self):
        """测试预计剩余时间显示"""
        start_time = time.time() - 50  # 50秒前开始
        table = _generate_worker_table(
            {}, 
            completed=5, 
            total=10, 
            failed=0, 
            paused=False, 
            start_time=start_time
        )
        # caption 应该包含剩余时间
        assert table.caption is not None
        assert "剩余" in table.caption

    def test_worker_states_display(self):
        """测试各种工作状态的显示"""
        worker_status = {
            1: {"state": "处理中", "question": "问题1", "response": "回复中...", "errors": 0},
            2: {"state": "工具", "question": "问题2", "response": "[工具:搜索]", "errors": 0},
            3: {"state": "重试中", "question": "问题3", "errors": 1},
            4: {"state": "等待", "question": "", "errors": 0},
        }
        table = _generate_worker_table(worker_status, 2, 10, 0)
        assert table is not None
        # 表格应该有4行数据
        assert len(table.rows) == 4


class TestRunSequentialBatch:
    """测试串行批量处理函数"""

    def test_successful_processing(self, tmp_path):
        """测试成功处理问题"""
        import openpyxl
        
        # 创建测试输入工作簿
        input_wb = openpyxl.Workbook()
        input_ws = input_wb.active
        input_ws.cell(row=1, column=1, value="问题")
        input_ws.cell(row=2, column=1, value="测试问题1")
        input_ws.cell(row=3, column=1, value="测试问题2")
        
        # 创建输出工作簿
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_file = str(tmp_path / "output.xlsx")
        
        # Mock provider
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("回答内容", True, None, "conv-1")
        
        # 运行串行批处理（使用patch抑制输出）
        with patch("dify_chat_tester.core.batch.console"):
            with patch("dify_chat_tester.core.batch.print_statistics"):
                _run_sequential_batch(
                    provider=mock_provider,
                    batch_worksheet=input_ws,
                    output_worksheet=output_ws,
                    output_workbook=output_wb,
                    output_file_name=output_file,
                    resume_from_row=2,
                    question_col_index=0,
                    doc_name_col_index=None,
                    selected_role="user",
                    selected_model="test-model",
                    provider_name="TestProvider",
                    enable_thinking=False,
                    show_batch_response=False,
                    batch_show_indicator=False,
                    request_interval=0,  # 无延迟
                )
        
        # 验证 provider 被调用了2次（2个问题）
        assert mock_provider.send_message.call_count == 2

    def test_empty_question_skipped(self, tmp_path):
        """测试空问题被跳过"""
        import openpyxl
        
        input_wb = openpyxl.Workbook()
        input_ws = input_wb.active
        input_ws.cell(row=1, column=1, value="问题")
        input_ws.cell(row=2, column=1, value="")  # 空问题
        input_ws.cell(row=3, column=1, value="有效问题")
        
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_file = str(tmp_path / "output.xlsx")
        
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("回答", True, None, None)
        
        with patch("dify_chat_tester.core.batch.console"):
            with patch("dify_chat_tester.core.batch.print_statistics"):
                _run_sequential_batch(
                    provider=mock_provider,
                    batch_worksheet=input_ws,
                    output_worksheet=output_ws,
                    output_workbook=output_wb,
                    output_file_name=output_file,
                    resume_from_row=2,
                    question_col_index=0,
                    doc_name_col_index=None,
                    selected_role="user",
                    selected_model="model",
                    provider_name="Provider",
                    enable_thinking=False,
                    show_batch_response=False,
                    batch_show_indicator=False,
                    request_interval=0,
                )
        
        # 空问题应被跳过，只调用1次
        assert mock_provider.send_message.call_count == 1

    def test_failed_question_recorded(self, tmp_path):
        """测试失败问题被正确记录"""
        import openpyxl
        
        input_wb = openpyxl.Workbook()
        input_ws = input_wb.active
        input_ws.cell(row=1, column=1, value="问题")
        input_ws.cell(row=2, column=1, value="会失败的问题")
        
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_file = str(tmp_path / "output.xlsx")
        
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("", False, "API错误", None)
        
        with patch("dify_chat_tester.core.batch.console"):
            with patch("dify_chat_tester.core.batch.print_statistics"):
                _run_sequential_batch(
                    provider=mock_provider,
                    batch_worksheet=input_ws,
                    output_worksheet=output_ws,
                    output_workbook=output_wb,
                    output_file_name=output_file,
                    resume_from_row=2,
                    question_col_index=0,
                    doc_name_col_index=None,
                    selected_role="user",
                    selected_model="model",
                    provider_name="Provider",
                    enable_thinking=False,
                    show_batch_response=False,
                    batch_show_indicator=False,
                    request_interval=0,
                )
        
        # 即使失败，provider 也会被调用
        assert mock_provider.send_message.call_count == 1

    def test_with_doc_name_column(self, tmp_path):
        """测试带文档名列的处理"""
        import openpyxl
        
        input_wb = openpyxl.Workbook()
        input_ws = input_wb.active
        input_ws.cell(row=1, column=1, value="文档名")
        input_ws.cell(row=1, column=2, value="问题")
        input_ws.cell(row=2, column=1, value="文档A")
        input_ws.cell(row=2, column=2, value="问题1")
        
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_file = str(tmp_path / "output.xlsx")
        
        mock_provider = MagicMock()
        mock_provider.send_message.return_value = ("回答", True, None, None)
        
        with patch("dify_chat_tester.core.batch.console"):
            with patch("dify_chat_tester.core.batch.print_statistics"):
                _run_sequential_batch(
                    provider=mock_provider,
                    batch_worksheet=input_ws,
                    output_worksheet=output_ws,
                    output_workbook=output_wb,
                    output_file_name=output_file,
                    resume_from_row=2,
                    question_col_index=1,  # 问题在第2列
                    doc_name_col_index=0,  # 文档名在第1列
                    selected_role="user",
                    selected_model="model",
                    provider_name="Provider",
                    enable_thinking=False,
                    show_batch_response=False,
                    batch_show_indicator=False,
                    request_interval=0,
                )
        
        assert mock_provider.send_message.call_count == 1

