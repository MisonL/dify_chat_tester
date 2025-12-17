"""
批量查询管理模块
负责处理批量询问模式的功能
"""

import json
import os
import time
import threading
import sys
import select
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import openpyxl

from dify_chat_tester.cli.terminal import (
    Panel,
    Text,
    box,
    console,
    print_error,
    print_file_list,
    print_input_prompt,
    print_statistics,
    print_success,
    print_warning,
)
from dify_chat_tester.config.loader import get_config
from dify_chat_tester.utils.excel import init_excel_log, log_to_excel

# Rich 组件用于并发显示
from rich.live import Live
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, MofNCompleteColumn, TimeRemainingColumn
from concurrent.futures import wait, FIRST_COMPLETED

# 从配置中获取批量保存间隔，默认每 10 条保存一次
_config = get_config()
SAVE_EVERY_N_QUERIES = _config.get_int("BATCH_SAVE_INTERVAL", 10) if _config else 10


def wait_for_any(futures: set, timeout: float = None):
    """等待任意一个 future 完成，返回 (已完成集合, 未完成集合)"""
    if not futures:
        return set(), set()
    done, not_done = wait(futures, timeout=timeout, return_when=FIRST_COMPLETED)
    return done, not_done


class KeyboardControl:
    """键盘控制类，用于在并发处理期间检测用户按键"""
    
    def __init__(self):
        self.stop_requested = False
        self.paused = False
        self._listener_thread = None
        self._running = False
    
    def start(self):
        """启动键盘监听"""
        self._running = True
        self._listener_thread = threading.Thread(target=self._listen, daemon=True)
        self._listener_thread.start()
    
    def stop(self):
        """停止键盘监听"""
        self._running = False
    
    def _listen(self):
        """后台监听键盘输入"""
        import tty
        import termios
        
        old_settings = termios.tcgetattr(sys.stdin)
        try:
            tty.setcbreak(sys.stdin.fileno())
            while self._running:
                if select.select([sys.stdin], [], [], 0.1)[0]:
                    ch = sys.stdin.read(1).lower()
                    if ch == 'q':
                        self.stop_requested = True
                    elif ch == 'p':
                        self.paused = not self.paused
        except Exception:
            pass  # 忽略终端不支持的情况
        finally:
            termios.tcsetattr(sys.stdin, termios.TCSADRAIN, old_settings)


def _run_sequential_batch(
    provider,
    batch_worksheet,
    output_worksheet,
    output_workbook,
    output_file_name,
    resume_from_row,
    question_col_index,
    doc_name_col_index,
    selected_role,
    selected_model,
    provider_name,
    enable_thinking,
    show_batch_response,
    batch_show_indicator,
    request_interval,
):
    """运行串行批量处理逻辑（封装了原有的批量处理核心循环）"""
    total_queries = 0
    successful_queries = 0
    failed_queries = 0
    queries_since_last_save = 0
    start_time = time.time()
    total_rows = batch_worksheet.max_row - 1

    # 获取列名用于统计显示
    column_names = [cell.value for cell in batch_worksheet[1]]

    try:
        for row_idx in range(
            resume_from_row, batch_worksheet.max_row + 1
        ):  # 从指定行开始读取数据
            # 获取文档名称（如果输入表中存在对应列）
            doc_name = ""
            if doc_name_col_index is not None:
                doc_cell_value = batch_worksheet.cell(
                    row=row_idx, column=doc_name_col_index + 1
                ).value
                doc_name = str(doc_cell_value) if doc_cell_value is not None else ""

            question_cell_value = batch_worksheet.cell(
                row=row_idx, column=question_col_index + 1
            ).value
            question = (
                str(question_cell_value) if question_cell_value is not None else ""
            )  # 确保转换为字符串

            if not question.strip():  # 检查问题是否为空或只包含空格
                print(f"警告: 第 {row_idx} 行问题为空，跳过。", file=console.file)
                failed_queries += 1  # 空问题也算作失败
                log_to_excel(
                    output_worksheet,
                    [
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        selected_role,
                        doc_name,
                        question,  # 原始问题为空
                        "",
                        False,
                        "问题为空",
                        0,
                        "",
                    ],
                )
                continue  # 跳过当前循环的剩余部分

            total_queries += 1  # 只有非空问题才计入总数

            # 计算进度
            current_progress = row_idx - 1
            pending_count = total_rows - current_progress
            progress_percent = (current_progress / total_rows) * 100

            # 美化问题显示（加粗和颜色）
            question_display = (
                f"[bold bright_magenta]处理进度 ({current_progress}/{total_rows} - {progress_percent:.1f}%) "
                f"| 待处理: {pending_count} | 问题:[/bold bright_magenta] "
                f"[bold yellow]{question[:50]}{'...' if len(question) > 50 else ''}[/bold yellow]"
            )
            console.print(f"\n{question_display}")

            response, success, error, conversation_id = provider.send_message(
                message=question,
                model=selected_model,
                role=selected_role,
                stream=True,
                show_indicator=batch_show_indicator,
                show_thinking=enable_thinking,
            )

            if success:
                successful_queries += 1
                # 如果开启了显示响应，在流式结束后（因 transient=True 会消失），需要重新打印一次最终结果使其保留在屏幕上
                if show_batch_response:
                    console.print(
                        Panel(
                            response,
                            title=f"🤖 {provider_name} 最终响应",
                            border_style="green",
                            box=box.ROUNDED,
                            padding=(0, 2),
                        )
                    )
                print(f"问题 (第 {total_queries} 个) 处理完成。")  # 简洁提示
            else:
                failed_queries += 1
                print(
                    f"问题 (第 {total_queries} 个) 处理失败。错误: {error}"
                )  # 简洁提示

            # 记录详细日志到日志文件
            log_to_excel(
                output_worksheet,
                [
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    selected_role,
                    doc_name,
                    question,
                    response,
                    success,
                    error,
                    conversation_id or "",
                ],
            )

            # 按批次保存日志，减少磁盘 IO
            queries_since_last_save += 1
            if queries_since_last_save >= SAVE_EVERY_N_QUERIES:
                try:
                    output_workbook.save(output_file_name)
                    queries_since_last_save = 0
                except PermissionError:
                    print_error(
                        f"警告：无法保存日志文件 '{output_file_name}'。请确保文件未被其他程序打开。"
                    )
                except Exception as e:
                    print_error(f"警告：保存日志时出错：{e}")

            time.sleep(request_interval)  # 间隔时间

    except KeyboardInterrupt:
        print_warning("\n⚠️  用户中断批量处理。正在保存当前进度...")
        try:
            output_workbook.save(output_file_name)
            print_success(f"进度已保存到: {output_file_name}")
        except Exception as e:
            print_error(f"保存进度失败: {e}")
        raise

    # 循环结束后做一次最终保存
    try:
        output_workbook.save(output_file_name)
    except PermissionError:
        print_error(
            f"警告：无法保存日志文件 '{output_file_name}'。请确保文件未被其他程序打开。"
        )
    except Exception as e:
        print_error(f"警告：保存日志时出错：{e}")

    end_time = time.time()
    total_duration = end_time - start_time

    # 统计信息面板
    print_statistics(total_queries, successful_queries, failed_queries, total_duration)

    # ----------------------------------------
    # 如果需要在函数内打印统计汇总信息，可以复用之前逻辑
    # 也可以将统计数据返回给外层调用
    # 为了保持逻辑完整性，这里直接打印
    # ----------------------------------------

    # 执行信息汇总面板
    summary_text = Text()
    summary_text.append("📁 文件信息\n", style="bold yellow")
    # 注意：Excel 文件名未传入，这里需要传递或者省略
    # 由于是 helper 函数，可以简化或通过参数传入文件名
    # 这里我们简化处理，不再重新打印文件路径，因为外层已经打印过了，或者需要再传入 selected_excel_file
    # 不过为了体验一致，我们还是尽量补全信息
    summary_text.append(f"  • 日志文件: {output_file_name} (自动关联)\n\n", style="white")

    summary_text.append("🤖 模型配置\n", style="bold yellow")
    summary_text.append(f"  • AI 供应商: {provider_name}\n", style="white")
    summary_text.append(f"  • 选用模型: {selected_model}\n", style="white")
    summary_text.append(f"  • 角色设定: {selected_role}\n", style="white")

    # 添加 API 接口地址
    base_url = getattr(provider, "base_url", None)
    if base_url:
        summary_text.append(f"  • API 接口: {base_url}\n", style="white")

    summary_text.append("\n", style="white")

    success_rate = (
        (successful_queries / total_queries * 100) if total_queries > 0 else 0
    )
    summary_text.append("📊 执行统计\n", style="bold yellow")
    summary_text.append(f"  • 成功率: {success_rate:.1f}%\n", style="white")
    summary_text.append(f"  • 请求间隔: {request_interval}秒\n", style="white")

    summary_panel = Panel(
        summary_text,
        title="[bold]📋 执行信息汇总[/bold]",
        border_style="bright_magenta",
        box=box.ROUNDED,
        padding=(1, 2),
    )
    console.print()
    console.print(summary_panel)
    console.print()


def _process_single_question(
    provider,
    question: str,
    selected_model: str,
    selected_role: str,
    enable_thinking: bool,
    worker_status: dict = None,
    worker_id: int = None,
):
    """处理单个问题的任务函数"""
    # 创建流式回调（如果提供了 worker_status）
    stream_callback = None
    if worker_status is not None and worker_id is not None:
        def stream_callback(event_type, content):
            """流式回调更新 worker_status"""
            if event_type == "text":
                # 显示回复的最后部分
                preview = content[-35:] if len(content) > 35 else content
                worker_status[worker_id]["response"] = preview
            elif event_type == "tool_call":
                worker_status[worker_id]["response"] = f"[工具:{content}]"
                worker_status[worker_id]["state"] = "工具"
            elif event_type == "thinking":
                worker_status[worker_id]["response"] = "[思考中...]"
    
    return provider.send_message(
        message=question,
        model=selected_model,
        role=selected_role,
        stream=True,
        show_indicator=False,  # 后台执行时不显示加载指示器
        show_thinking=enable_thinking,
        stream_callback=stream_callback,
    )


def _process_with_retry(
    provider,
    question: str,
    selected_model: str,
    selected_role: str,
    enable_thinking: bool,
    max_retries: int = 3,
    worker_status: dict = None,
    worker_id: int = None,
):
    """带重试的问题处理函数，最多重试 max_retries 次"""
    last_error = None
    retry_count = 0
    
    for attempt in range(max_retries + 1):
        try:
            result = _process_single_question(
                provider, question, selected_model, selected_role, enable_thinking,
                worker_status, worker_id
            )
            response, success, error, conversation_id = result
            
            if success:
                return result, retry_count
            else:
                # API 返回失败但没有异常
                last_error = error
                retry_count += 1
                if attempt < max_retries:
                    time.sleep(1)  # 重试前等待 1 秒
                    continue
        except Exception as e:
            last_error = str(e)
            retry_count += 1
            if attempt < max_retries:
                time.sleep(1)
                continue
    
    # 所有重试都失败
    return ("", False, f"重试{max_retries}次后失败: {last_error}", None), retry_count


def _generate_worker_table(
    worker_status: dict,
    completed: int,
    total: int,
    failed: int,
    paused: bool = False,
    start_time: float = None,
) -> Table:
    """生成工作线程状态表格"""
    # 计算进度百分比
    percent = (completed / total * 100) if total > 0 else 0
    
    # 计算预计剩余时间
    eta_text = ""
    if start_time and completed > 0:
        elapsed = time.time() - start_time
        avg_time = elapsed / completed
        remaining = (total - completed) * avg_time
        if remaining > 3600:
            eta_text = f"{remaining/3600:.1f}h"
        elif remaining > 60:
            eta_text = f"{remaining/60:.1f}m"
        else:
            eta_text = f"{remaining:.0f}s"
    
    # 构建标题（优化间距）
    if paused:
        status_text = "[bold yellow]⏸ 已暂停[/bold yellow]"
    else:
        status_text = f"[bold cyan]{completed}[/bold cyan]/[dim]{total}[/dim]"
    
    title = f"📊 并发处理  {status_text}  ✅ {completed-failed}  ❌ {failed}  [dim](P=暂停 Q=停止 Ctrl+C=退出)[/dim]"
    
    # 构建进度条
    bar_width = 40
    filled = int(bar_width * percent / 100)
    bar = "█" * filled + "░" * (bar_width - filled)
    eta_display = f"  预计剩余: {eta_text}" if eta_text else ""
    caption = f"[cyan]{bar}[/cyan]  [bold]{percent:.1f}%[/bold]{eta_display}"
    
    table = Table(title=title, caption=caption, box=box.ROUNDED)
    table.add_column("线程", style="cyan", width=6)
    table.add_column("状态", style="green", width=10)
    table.add_column("错误", style="red", width=4, justify="center")
    table.add_column("回复预览", style="yellow", max_width=45)
    
    for worker_id, status in sorted(worker_status.items()):
        state = status.get("state", "空闲")
        question = status.get("question", "")
        response = status.get("response", "")
        error_count = status.get("errors", 0)
        
        # 根据状态显示不同内容
        if state == "处理中" and response:
            preview = response
        elif state == "工具":
            preview = response if response else "[调用工具...]"
        elif state in ["完成", "失败"]:
            preview = response if response else question[:30] + "..."
        else:
            # 等待/初始状态显示问题
            preview = question[:35] + "..." if len(question) > 35 else question
        
        if state == "处理中":
            state_display = "[bold cyan]🔄 处理中[/bold cyan]"
        elif state == "完成":
            state_display = "[bold green]✅ 完成[/bold green]"
        elif state == "失败":
            state_display = "[bold red]❌ 失败[/bold red]"
        elif state == "重试中":
            state_display = "[bold yellow]🔁 重试中[/bold yellow]"
        elif state == "工具":
            state_display = "[bold magenta]🔧 工具[/bold magenta]"
        else:
            state_display = "[dim]⏳ 等待[/dim]"
        
        # 错误数显示
        error_display = f"[red]{error_count}[/red]" if error_count > 0 else "[dim]0[/dim]"
        
        table.add_row(f"#{worker_id}", state_display, error_display, preview)
    
    return table

def _run_concurrent_batch(
    provider,
    batch_worksheet,
    output_worksheet,
    output_workbook,
    output_file_name,
    resume_from_row,
    question_col_index,
    doc_name_col_index,
    selected_role,
    selected_model,
    provider_name,
    enable_thinking,
    show_batch_response,
    concurrency,
):
    """运行并发批量处理逻辑"""

    total_queries = 0
    successful_queries = 0
    failed_queries = 0
    queries_since_last_save = 0
    start_time = time.time()
    total_rows = batch_worksheet.max_row - 1

    # 获取列名用于统计显示
    column_names = [cell.value for cell in batch_worksheet[1]]

    # 准备任务队列
    tasks = []
    console.print(f"\n[bold cyan]🚀 已启动并发模式 (并发数: {concurrency})[/bold cyan]")
    
    # 预读取所有待处理的问题
    for row_idx in range(resume_from_row, batch_worksheet.max_row + 1):
        doc_name = ""
        if doc_name_col_index is not None:
            doc_cell_value = batch_worksheet.cell(
                row=row_idx, column=doc_name_col_index + 1
            ).value
            doc_name = str(doc_cell_value) if doc_cell_value is not None else ""

        question_cell_value = batch_worksheet.cell(
            row=row_idx, column=question_col_index + 1
        ).value
        question = (
            str(question_cell_value) if question_cell_value is not None else ""
        )
        
        tasks.append({
            "row_idx": row_idx,
            "doc_name": doc_name,
            "question": question,
            "index": len(tasks)  # 相对索引，用于结果排序
        })

    if not tasks:
        print_success("没有需要处理的任务。")
        return

    # 结果缓冲区 {index: result_tuple}
    results_buffer = {}
    # 工作线程状态追踪 {worker_id: {"state": "处理中/完成/失败", "question": "..."}}
    worker_status = {i: {"state": "等待", "question": ""} for i in range(1, concurrency + 1)}
    completed_count = 0
    failed_count = 0
    total_tasks = len(tasks)
    
    # 启动键盘控制
    kb_control = KeyboardControl()
    kb_control.start()
    user_stopped = False  # 用户主动停止标志

    try:
        with Live(console=console, refresh_per_second=4) as live:
            with ThreadPoolExecutor(max_workers=concurrency) as executor:
                # 提交任务字典 {future: (task_info, worker_id)}
                future_to_task = {}
                pending_tasks = list(tasks)  # 待提交的任务队列
                active_futures = set()  # 当前活跃的 future
                next_worker_id = 1  # 下一个可用的 worker ID
                
                # 初始提交 concurrency 个任务
                while pending_tasks and len(active_futures) < concurrency:
                    task = pending_tasks.pop(0)
                    if not task["question"].strip():
                        # 空问题直接标记为完成
                        results_buffer[task["index"]] = ("", False, "问题为空", None)
                        completed_count += 1
                        failed_count += 1
                        continue
                    
                    worker_id = next_worker_id
                    next_worker_id = (next_worker_id % concurrency) + 1
                    
                    worker_status[worker_id] = {"state": "处理中", "question": task["question"], "errors": 0}
                    
                    future = executor.submit(
                        _process_with_retry,
                        provider,
                        task["question"],
                        selected_model,
                        selected_role,
                        enable_thinking,
                        3,  # max_retries
                        worker_status,  # 传递 worker_status
                        worker_id,  # 传递 worker_id
                    )
                    future_to_task[future] = (task, worker_id)
                    active_futures.add(future)
                
                # 更新显示
                live.update(_generate_worker_table(worker_status, completed_count, total_tasks, failed_count, kb_control.paused, start_time))
                
                # 处理完成的任务并提交新任务
                while active_futures or pending_tasks:
                    # 检查用户是否请求停止
                    if kb_control.stop_requested:
                        user_stopped = True
                        print_warning("\n⚠️ 用户请求停止，正在等待当前任务完成...")
                        break
                    
                    # 如果暂停，只更新显示，不处理新任务
                    if kb_control.paused:
                        # 首次进入暂停状态时提示
                        if not getattr(kb_control, '_pause_notified', False):
                            console.print("\n[bold yellow]⏸ 已暂停 - 按 P 恢复，按 Q 保存并停止[/bold yellow]")
                            kb_control._pause_notified = True
                        live.update(_generate_worker_table(worker_status, completed_count, total_tasks, failed_count, True, start_time))
                        time.sleep(0.3)
                        continue
                    else:
                        # 从暂停恢复时打印提示
                        if getattr(kb_control, '_pause_notified', False):
                            console.print("\n[bold green]▶ 已恢复处理[/bold green]")
                        kb_control._pause_notified = False  # 重置暂停通知状态
                    
                    # 等待任意一个任务完成
                    done, active_futures = wait_for_any(active_futures, timeout=0.5)
                    
                    for future in done:
                        task, worker_id = future_to_task[future]
                        try:
                            future_result = future.result()
                            # _process_with_retry 返回 (result, retry_count)
                            result, retry_count = future_result
                        except Exception as e:
                            result = ("", False, str(e), None)
                            retry_count = 0
                        
                        results_buffer[task["index"]] = result
                        completed_count += 1
                        
                        # 更新状态和错误计数
                        current_errors = worker_status.get(worker_id, {}).get("errors", 0) + retry_count
                        success = result[1] if len(result) > 1 else False
                        if success:
                            worker_status[worker_id] = {"state": "完成", "question": task["question"], "errors": current_errors}
                        else:
                            worker_status[worker_id] = {"state": "失败", "question": task["question"], "errors": current_errors}
                            failed_count += 1
                        
                        # 提交下一个任务（如果有）
                        while pending_tasks:
                            next_task = pending_tasks.pop(0)
                            if not next_task["question"].strip():
                                results_buffer[next_task["index"]] = ("", False, "问题为空", None)
                                completed_count += 1
                                failed_count += 1
                                continue
                            
                            # 保留原有错误计数
                            prev_errors = worker_status.get(worker_id, {}).get("errors", 0)
                            worker_status[worker_id] = {"state": "处理中", "question": next_task["question"], "errors": prev_errors}
                            
                            new_future = executor.submit(
                                _process_with_retry,
                                provider,
                                next_task["question"],
                                selected_model,
                                selected_role,
                                enable_thinking,
                                3,  # max_retries
                                worker_status,  # 传递 worker_status
                                worker_id,  # 传递 worker_id
                            )
                            future_to_task[new_future] = (next_task, worker_id)
                            active_futures.add(new_future)
                            break
                        else:
                            # 没有更多任务，将 worker 标记为空闲
                            if worker_id in worker_status:
                                old_q = worker_status[worker_id].get("question", "")
                                old_errors = worker_status[worker_id].get("errors", 0)
                                worker_status[worker_id] = {"state": "完成", "question": old_q, "errors": old_errors}
                    
                    # 更新显示
                    live.update(_generate_worker_table(worker_status, completed_count, total_tasks, failed_count, kb_control.paused, start_time))

    except KeyboardInterrupt:
        kb_control.stop()
        print_warning("\n⚠️  用户中断批量处理 (Ctrl+C)。正在保存当前进度...")
        # 此时 executor 会尝试 join，可能需要一段时间
        raise
    finally:
        kb_control.stop()
    
    # 收集所有失败的任务进行批量重试
    failed_tasks = []
    for task in tasks:
        idx = task["index"]
        if idx in results_buffer:
            result = results_buffer[idx]
            if not result[1]:  # success == False
                failed_tasks.append(task)
    
    # 如果有失败任务且用户没有主动停止，进行批量重试
    if failed_tasks and not user_stopped:
        console.print(f"\n[bold yellow]🔄 发现 {len(failed_tasks)} 个失败任务，开始批量重试...[/bold yellow]")
        
        retry_success = 0
        retry_failed = 0
        
        with ThreadPoolExecutor(max_workers=concurrency) as retry_executor:
            retry_futures = {}
            for task in failed_tasks:
                future = retry_executor.submit(
                    _process_with_retry,
                    provider,
                    task["question"],
                    selected_model,
                    selected_role,
                    enable_thinking,
                    3,  # max_retries
                )
                retry_futures[future] = task
            
            for future in as_completed(retry_futures):
                task = retry_futures[future]
                try:
                    result, _ = future.result()
                except Exception as e:
                    result = ("", False, str(e), None)
                
                # 更新结果缓冲区
                results_buffer[task["index"]] = result
                
                if result[1]:  # success
                    retry_success += 1
                else:
                    retry_failed += 1
        
        console.print(f"[bold green]✅ 批量重试完成: 成功 {retry_success}, 仍失败 {retry_failed}[/bold green]")
    
    # 处理结果
    if user_stopped:
        console.print("\n[bold yellow]⚠️ 用户请求停止，部分任务未完成。正在保存已完成的结果...[/bold yellow]")
    else:
        console.print("\n[bold green]✅ 所有请求处理完成，正在写入结果...[/bold green]")
    
    for task in tasks:
        idx = task["index"]
        if idx not in results_buffer:
             # 空问题等情况已经在循环前处理了，或者异常丢失
             if not task["question"].strip():
                 result = ("", False, "问题为空", None)
             else:
                 result = ("", False, "任务未完成或丢失", None)
        else:
            result = results_buffer[idx]
            
        response, success, error, conversation_id = result
        
        # 统计
        if not task["question"].strip():
             failed_queries += 1
        else:
            total_queries += 1
            if success:
                successful_queries += 1
            else:
                failed_queries += 1

        # 显示（可选，如果用户开启了显示响应）
        # 并发模式下，我们在最后统一显示可能会刷屏，或者只显示失败的？
        # 设计方案中提到“顺序流式输出”，这里简化为“顺序显示结果”
        if success and show_batch_response:
             console.print(f"\n[bold yellow]Q ({task['row_idx']}): {task['question']}[/bold yellow]")
             console.print(Panel(response, title=f"A: {provider_name}", border_style="green"))
        elif not success and task["question"].strip():
             console.print(f"\n[bold red]Q ({task['row_idx']}): {task['question']} - 失败: {error}[/bold red]")

        # 写入 Excel
        log_to_excel(
            output_worksheet,
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                selected_role,
                task["doc_name"],
                task["question"],
                response,
                success,
                error,
                conversation_id or "",
            ],
        )

        # 批量保存
        queries_since_last_save += 1
        if queries_since_last_save >= SAVE_EVERY_N_QUERIES:
            try:
                output_workbook.save(output_file_name)
                queries_since_last_save = 0
            except Exception as e:
                print_error(f"警告：保存日志时出错：{e}")

    # 最终保存
    try:
        output_workbook.save(output_file_name)
    except Exception as e:
        print_error(f"警告：保存日志时出错：{e}")

    end_time = time.time()
    total_duration = end_time - start_time

    # 打印统计
    print_statistics(total_queries, successful_queries, failed_queries, total_duration)
    
    # 汇总信息（复用部分逻辑，从简）
    print_success(f"并发批量处理完成。日志已保存至: {output_file_name}")


def run_batch_query(
    provider,
    selected_role: str,
    provider_name: str,
    selected_model: str,
    batch_request_interval: float,
    batch_default_show_response: bool,
    concurrency: int = 1,
):
    """运行批量询问模式
    
    Args:
        provider: AI 提供商实例
        selected_role: 角色名称
        provider_name: 提供商名称
        selected_model: 模型名称
        batch_request_interval: 请求间隔时间（秒）
        batch_default_show_response: 是否默认显示响应
        concurrency: 并发数（1=串行，2-10=并发）
    """
    # 获取配置
    config = get_config()
    enable_thinking = config.get_enable_thinking()
    console.print()

    # 模式信息面板
    mode_text = Text()
    mode_text.append("🤖 模型: ", style="bold yellow")
    mode_text.append(f"{selected_model}\n", style="bold cyan")
    mode_text.append("👤 角色: ", style="bold yellow")
    mode_text.append(f"{selected_role}\n", style="bold cyan")
    mode_text.append("💬 供应商: ", style="bold yellow")
    mode_text.append(f"{provider_name}", style="bold cyan")

    # Dify 不再需要显示应用 ID

    mode_panel = Panel(
        mode_text,
        title="[bold]📄 批量询问模式[/bold]",
        border_style="bright_magenta",
        box=box.ROUNDED,
        padding=(1, 2),
    )
    console.print(mode_panel)
    console.print()

    # 列出当前目录下的 Excel 文件
    excel_files = [
        f for f in os.listdir(".") if f.endswith(".xlsx") and os.path.isfile(f)
    ]

    selected_excel_file = None
    while True:
        if excel_files:
            print_file_list(excel_files)
            file_input = print_input_prompt("请输入 Excel 文件序号或直接输入文件路径")

            try:
                file_index = int(file_input)
                if 1 <= file_index <= len(excel_files):
                    excel_file_path = excel_files[file_index - 1]
                else:
                    print(
                        f"错误: 无效的文件序号 '{file_index}'。请重新输入。",
                        file=console.file,
                    )
                    continue
            except ValueError:
                # 用户输入的是路径
                excel_file_path = file_input
        else:
            excel_file_path = print_input_prompt(
                "当前目录下没有找到 Excel 文件，请输入包含询问内容的 Excel 文件路径"
            )

        if not os.path.exists(excel_file_path):
            print(
                f"错误: 文件 '{excel_file_path}' 不存在。请重新输入。",
                file=console.file,
            )
            continue

        try:
            batch_workbook = openpyxl.load_workbook(excel_file_path)
            batch_worksheet = batch_workbook.active
            if batch_worksheet is None:  # 确保工作表不为None
                print(
                    f"错误: Excel 文件 '{excel_file_path}' 中没有活动工作表。请重新输入。",
                    file=console.file,
                )
                continue
            selected_excel_file = excel_file_path
            break  # 成功读取文件并获取工作表，跳出循环
        except Exception as e:
            print(
                f"错误: 无法读取 Excel 文件 '{excel_file_path}'。请确保文件格式正确且未被占用。错误信息: {e}。请重新输入。",
                file=console.file,
            )
            continue

    # 检测用户是否选择了日志文件（_log.xlsx 结尾）
    # 如果是，尝试找到对应的原始输入文件并提供恢复选项
    input_dir = os.path.dirname(selected_excel_file) or "."
    input_basename = os.path.splitext(os.path.basename(selected_excel_file))[0]
    
    if input_basename.endswith("_log"):
        # 用户选择的是日志文件，尝试找到原始输入文件
        original_basename = input_basename[:-4]  # 移除 "_log" 后缀
        original_file_path = os.path.join(input_dir, f"{original_basename}.xlsx")
        
        if os.path.exists(original_file_path):
            console.print(
                Panel(
                    f"检测到您选择的是日志文件: [bold cyan]{selected_excel_file}[/bold cyan]\n"
                    f"找到对应的原始输入文件: [bold green]{original_file_path}[/bold green]\n\n"
                    f"系统将使用原始输入文件继续处理，并根据日志文件自动恢复进度。",
                    title="[bold yellow]📋 智能恢复检测[/bold yellow]",
                    border_style="yellow",
                    box=box.ROUNDED,
                )
            )
            
            use_original = (
                print_input_prompt(
                    f"是否使用原始文件 '{original_file_path}' 继续处理？(Y/n)"
                )
                .strip()
                .lower()
            )
            
            if not use_original or use_original in ("y", "yes"):
                # 重新加载原始文件
                try:
                    batch_workbook = openpyxl.load_workbook(original_file_path)
                    batch_worksheet = batch_workbook.active
                    if batch_worksheet is None:
                        print_error("原始文件没有活动工作表，将继续使用日志文件。")
                    else:
                        selected_excel_file = original_file_path
                        # 重新计算 basename
                        input_basename = original_basename
                        print_success(f"已切换到原始输入文件: {original_file_path}")
                except Exception as e:
                    print_error(f"无法加载原始文件: {e}，将继续使用日志文件。")
        else:
            print_warning(
                f"您选择的是日志文件，但未找到对应的原始输入文件 '{original_file_path}'。\n"
                f"将直接使用日志文件作为输入（可能导致重复处理已完成的内容）。"
            )

    # 为当前输入文件构建固定的日志文件路径
    # 规则：输入文件名（不含扩展名） + _log.xlsx
    # 注意：如果上面切换了文件，input_basename 已经更新
    input_dir = os.path.dirname(selected_excel_file) or "."
    input_basename = os.path.splitext(os.path.basename(selected_excel_file))[0]
    output_file_name = os.path.join(input_dir, f"{input_basename}_log.xlsx")

    # 默认从第二行开始（第一行为表头）
    resume_from_row = 2

    # 检测是否存在日志文件以判断进度
    if os.path.exists(output_file_name):
        try:
            # 尝试读取现有的日志文件
            existing_wb = openpyxl.load_workbook(output_file_name)
            existing_ws = existing_wb.active
            if existing_ws and existing_ws.max_row > 1:
                # 日志文件存在且有数据（不止表头）
                last_row = existing_ws.max_row
                # 理论上，日志行数 = 已处理行数 + 1 (表头)
                # 所以下一行输入行号 = 日志最大行号 + 1
                # 例如：日志有表头(1) + 1条数据(2) -> max_row=2 ->已处理1条 -> 下一条是输入文件的第3行
                # 验证：输入头(1) + 数据1(2). 输出头(1) + 数据1(2). resume = 2 + 1 = 3. 正确.
                potential_resume_row = last_row + 1

                if potential_resume_row <= batch_worksheet.max_row + 1:
                    processed_count = last_row - 1
                    console.print(
                        Panel(
                            f"检测到历史日志文件: [bold cyan]{output_file_name}[/bold cyan]\n"
                            f"已处理记录数: [bold green]{processed_count}[/bold green]\n"
                            f"上次结束位置: 第 {last_row} 行 (对应输入文件第 {potential_resume_row-1} 行)",
                            title="[bold yellow]📋 恢复进度提示[/bold yellow]",
                            border_style="yellow",
                            box=box.ROUNDED,
                        )
                    )

                    resume_choice = (
                        print_input_prompt(
                            f"是否从第 {potential_resume_row} 行继续处理？(Y/n，选择 n 将覆盖旧日志)"
                        )
                        .strip()
                        .lower()
                    )

                    if not resume_choice or resume_choice in ("y", "yes"):
                        resume_from_row = potential_resume_row
                        print_success(f"已恢复进度，将从第 {resume_from_row} 行开始。")
                    else:
                        print_warning("已选择重新开始，旧的日志文件将被覆盖！")
                        resume_from_row = 2
        except Exception as e:
            print_error(f"读取现有日志文件失败: {e}，将重新开始。")
            resume_from_row = 2

    # 获取列名
    column_names = [cell.value for cell in batch_worksheet[1]]
    print_success(f"已选择文件: {selected_excel_file}")

    # 检查是否存在“文档名称”列
    doc_name_col_index = None
    try:
        doc_name_col_index = column_names.index("文档名称")
    except ValueError:
        doc_name_col_index = None

    # 让用户通过序号选择问题列
    from dify_chat_tester.cli.terminal import select_column_by_index

    question_col_index = select_column_by_index(column_names, "请选择问题所在列的序号")

    # 注意：不再创建或使用回答列，所有结果只记录到日志文件

    # 询问是否显示每个问题的回答内容（回车则使用配置中的默认值）
    display_response_choice = print_input_prompt(
        "是否在控制台显示每个问题的回答内容？ (Y/n，直接回车使用配置默认值)"
    )
    if not display_response_choice:
        show_batch_response = batch_default_show_response
    else:
        show_batch_response = display_response_choice.lower() != "n"

    # 询问并发数（如果未通过命令行指定）
    if concurrency <= 1:
        concurrency_input = print_input_prompt(
            "是否启用并发模式？(输入并发数 2-10，直接回车使用串行模式)"
        )
        if concurrency_input.strip():
            try:
                concurrency = int(concurrency_input)
                if concurrency < 1:
                    concurrency = 1
                elif concurrency > 10:
                    concurrency = 10
                    print_warning("并发数已限制为最大 10")
            except ValueError:
                concurrency = 1
    
    if concurrency > 1:
        print_success(f"已启用并发模式，并发数: {concurrency}")
    else:
        print_success("使用串行模式处理")

    # 从配置中获取请求间隔时间（使用配置中的默认值）
    request_interval = batch_request_interval

    # 为批量模式设置 show_indicator（只有在用户选择显示响应且为串行模式时才启用）
    # 并发模式下禁用逐条流式显示（会造成输出混乱）
    batch_show_indicator = show_batch_response and concurrency <= 1

    # 初始化日志文件（如果是恢复模式，init_excel_log 会加载现有文件；否则若不恢复，我们需要确保是覆盖还是追加？
    # init_excel_log逻辑是：存在则加载。
    # 如果我们选择了"不恢复"（resume_from_row=2），意味着我们想重写。
    # 所以如果 resume_from_row == 2 且文件存在，我们需要删除它以便 init_excel_log 创建新的（或者清空内容）。
    # 简单做法：如果 resume_from_row == 2，先尝试删除旧文件。
    if resume_from_row == 2 and os.path.exists(output_file_name):
        try:
            os.remove(output_file_name)
        except Exception:
            pass

    batch_log_headers = [
        "时间戳",
        "角色",
        "文档名称",
        "原始问题",
        f"{provider_name}响应",
        "是否成功",
        "错误信息",
        "对话ID",
    ]
    output_workbook, output_worksheet = init_excel_log(
        output_file_name, batch_log_headers
    )

    total_queries = 0
    successful_queries = 0
    failed_queries = 0
    # 自上次保存以来已处理的问题数量
    queries_since_last_save = 0
    start_time = time.time()

    total_rows = batch_worksheet.max_row - 1

    # 如果上次已经处理到文件末尾，则直接结束
    if resume_from_row > batch_worksheet.max_row:
        print_success("检测到该文件的所有问题均已处理完成，无需继续。")
        return

    print(
        f"\n开始批量询问... (共 {total_rows} 行数据，当前从第 {resume_from_row} 行开始)"
    )
    
    # 根据并发数选择处理模式
    if concurrency > 1:
        # 并发模式
        _run_concurrent_batch(
            provider=provider,
            batch_worksheet=batch_worksheet,
            output_worksheet=output_worksheet,
            output_workbook=output_workbook,
            output_file_name=output_file_name,
            resume_from_row=resume_from_row,
            question_col_index=question_col_index,
            doc_name_col_index=doc_name_col_index,
            selected_role=selected_role,
            selected_model=selected_model,
            provider_name=provider_name,
            enable_thinking=enable_thinking,
            show_batch_response=show_batch_response,
            concurrency=concurrency,
        )
    else:
        # 串行模式（原有逻辑）
        _run_sequential_batch(
            provider=provider,
            batch_worksheet=batch_worksheet,
            output_worksheet=output_worksheet,
            output_workbook=output_workbook,
            output_file_name=output_file_name,
            resume_from_row=resume_from_row,
            question_col_index=question_col_index,
            doc_name_col_index=doc_name_col_index,
            selected_role=selected_role,
            selected_model=selected_model,
            provider_name=provider_name,
            enable_thinking=enable_thinking,
            show_batch_response=show_batch_response,
            batch_show_indicator=batch_show_indicator,
            request_interval=request_interval,
        )
    return
