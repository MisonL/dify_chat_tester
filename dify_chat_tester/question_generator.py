"""
问题生成器模块
负责读取文档、生成测试问题，并导出到Excel
"""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from dify_chat_tester.terminal_ui import (
    console,
    print_error,
    print_info,
    print_success,
    print_warning,
)


def read_markdown_files(folder_path: str) -> Tuple[List[str], Dict[str, str]]:
    """
    读取指定文件夹中的所有MD文件

    Args:
        folder_path: 文件夹路径

    Returns:
        tuple: (文件名列表, 文件内容字典)
    """
    folder = Path(folder_path)

    if not folder.exists() or not folder.is_dir():
        print_error(f"文件夹不存在或不是有效目录: {folder_path}")
        return [], {}

    # 查找所有.md文件
    md_files = list(folder.glob("*.md"))

    if not md_files:
        print_warning(f"文件夹中没有找到MD文件: {folder_path}")
        return [], {}

    file_names = []
    file_contents = {}

    print_info(f"找到 {len(md_files)} 个MD文件")

    for md_file in md_files:
        try:
            with open(md_file, "r", encoding="utf-8") as f:
                content = f.read()
                file_names.append(md_file.name)
                file_contents[md_file.name] = content
                print(f"  ✓ {md_file.name} ({len(content)} 字符)")
        except Exception as e:
            print_warning(f"  ✗ 读取文件失败 {md_file.name}: {e}")
            continue

    console.print()
    return file_names, file_contents


# 单个提示中使用的最大文档内容长度（字符）
MAX_DOC_CHARS_PER_CALL = 8000


def generate_questions_for_document(
    provider, model: str, role: str, document_name: str, document_content: str
) -> List[str]:
    """
    为单个文档生成测试问题

    Args:
        provider: AI提供商实例
        model: 模型名称
        role: 角色
        document_name: 文档名称
        document_content: 文档内容

    Returns:
        list: 生成的问题列表
    """
    # 按长度对文档内容进行分块
    text = document_content or ""
    if not text:
        print_warning(f"文档 {document_name} 内容为空，跳过问题生成")
        return []

    chunks: List[str] = []
    for i in range(0, len(text), MAX_DOC_CHARS_PER_CALL):
        chunks.append(text[i : i + MAX_DOC_CHARS_PER_CALL])

    all_questions: List[str] = []
    total_chunks = len(chunks)

    print_info(f"正在为文档 {document_name} 生成问题（共 {total_chunks} 个分块）...")

    for idx, chunk in enumerate(chunks, start=1):
        # 从配置中获取提示词模板
        from dify_chat_tester.config_loader import get_config

        prompt_template = get_config().get_single_knowledge_prompt()

        # 构建提示词
        try:
            prompt = prompt_template.format(
                idx=idx,
                total_chunks=total_chunks,
                document_name=document_name,
                chunk=chunk,
            )
        except KeyError as e:
            # 如果用户自定义的模板缺少必要的占位符或包含未知占位符，回退到默认
            print_error(f"提示词模板格式错误: {e}，将使用默认模板")
            prompt = f"""你是一个专业的测试问题生成助手。请仔细阅读以下文档内容，生成尽可能多的测试问题。

当前为第 {idx}/{total_chunks} 个内容分块，请尽量覆盖本分块中的知识点。

要求：
1. 先识别本分块中有哪些独立的知识点或主题，并大致评估每个知识点被真实用户询问的概率（高/中/低）。
2. 问题需要模仿普通用户的真实提问语气，口语化、自然。
3. 问题应该覆盖文档中的各个知识点。
4. 问题的难度应该有所变化，包括简单查询、复杂推理等。
5. 每个问题都应该能从文档中找到答案依据。
6. 生成的问题数量不做固定限制：
   - 知识点较少时，覆盖所有知识点即可；
   - 对"高概率"知识点，多写几个不同角度的提问；
   - 对"中概率"知识点，至少生成 1-2 个提问；
   - 对"低概率"知识点，可适当精简。
7. 最终只输出问题列表，不要输出对知识点或概率的解释。
8. 以 JSON 数组格式返回，每个元素是一个问题字符串。

文档名称：{document_name}

文档内容（第 {idx}/{total_chunks} 块）：
{chunk}

请生成问题列表（必须是 JSON 数组格式，例如:["问题1","问题2","问题3"]）："""

        try:
            # 调用AI生成问题
            response, success, error_msg, _ = provider.send_message(
                message=prompt,
                model=model,
                role=role,
                stream=False,
                show_indicator=True,
            )

            if not success:
                print_error(f"AI调用失败（分块 {idx}/{total_chunks}）: {error_msg}")
                continue

            # 解析JSON响应
            questions = parse_questions_from_response(response)

            if questions:
                print_success(
                    f"文档 {document_name} - 分块 {idx}/{total_chunks} 生成 {len(questions)} 个问题"
                )
                all_questions.extend(questions)
            else:
                print_warning(
                    f"文档 {document_name} - 分块 {idx}/{total_chunks} 未能解析出问题"
                )
        except Exception as e:
            print_error(f"生成问题时发生错误（分块 {idx}/{total_chunks}）: {e}")
            continue

    # 去重并保持顺序
    seen = set()
    dedup_questions: List[str] = []
    for q in all_questions:
        q_norm = (q or "").strip()
        if not q_norm:
            continue
        if q_norm in seen:
            continue
        seen.add(q_norm)
        dedup_questions.append(q_norm)

    if dedup_questions:
        print_success(
            f"文档 {document_name} 共生成 {len(dedup_questions)} 个去重后的问题"
        )
    else:
        print_warning(f"文档 {document_name} 未生成任何有效问题")

    return dedup_questions


def parse_questions_from_response(response: str) -> List[str]:
    """
    从AI响应中解析问题列表

    Args:
        response: AI的响应文本

    Returns:
        list: 问题列表
    """
    questions = []

    if not response:
        return questions

    # 尝试直接解析JSON
    try:
        # 查找JSON数组
        start_idx = response.find("[")
        end_idx = response.rfind("]")

        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            json_str = response[start_idx : end_idx + 1].strip()
            if json_str:
                questions = json.loads(json_str)

                # 确保是字符串列表
                if isinstance(questions, list):
                    questions = [str(q).strip() for q in questions if q]
                    return questions
    except json.JSONDecodeError:
        # 保持静默，回退到按行解析
        pass

    # 如果JSON解析失败，尝试按行解析
    lines = response.split("\n")
    for raw_line in lines:
        line = raw_line.strip()
        # 跳过空行和JSON标记
        if not line or line in ["[", "]", "{", "}"]:
            continue

        # 记录原始是否为编号列表项或项目符号列表项
        is_numbered_item = bool(re.match(r"^\d+[\.\)]\s*\S+", line))
        is_bullet_item = bool(re.match(r"^[-*•\u2022]\s*\S+", line))

        # 尝试移除可能的列表前缀（如 "- ", "* ", "• ")
        line = re.sub(r"^[-*•\u2022]\s*", "", line)
        # 尝试移除序号前缀（如 "1. ", "2) " 等）
        line = re.sub(r"^\d+[\.\)]\s*", "", line)

        # 去掉首尾引号和逗号
        line = line.strip("\"'").strip(",").strip()

        # 只保留像问题的行：
        # - 以 ? 或 ？ 结尾，或
        # - 原始行是编号列表项，或
        # - 原始行为带项目符号的列表项
        if (
            line
            and len(line) > 1
            and (
                line.endswith("?")
                or line.endswith("？")
                or is_numbered_item
                or is_bullet_item
            )
        ):
            questions.append(line)

    return questions


def export_questions_to_excel(
    questions_data: List[Tuple[str, str]], output_path: str
) -> bool:
    """
    将问题数据导出到Excel

    Args:
        questions_data: [(文档名称, 问题), ...] 的列表
        output_path: 输出文件路径

    Returns:
        bool: 是否成功导出
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试问题"

        # 设置列名
        ws["A1"] = "文档名称"
        ws["B1"] = "问题"

        # 美化表头
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ["A1", "B1"]:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = header_alignment

        # 设置列宽
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 80

        # 写入数据
        for idx, (doc_name, question) in enumerate(questions_data, start=2):
            ws[f"A{idx}"] = doc_name
            ws[f"B{idx}"] = question

            # 设置文本对齐
            ws[f"A{idx}"].alignment = Alignment(vertical="top")
            ws[f"B{idx}"].alignment = Alignment(vertical="top", wrap_text=True)

        # 保存文件
        wb.save(output_path)
        return True

    except ImportError:
        print_error("缺少 openpyxl 库，无法导出Excel文件")
        return False
    except Exception as e:
        print_error(f"导出Excel时发生错误: {e}")
        return False


def run_question_generation(
    provider, role: str, provider_name: str, selected_model: str, folder_path: str
):
    """
    运行问题生成流程

    Args:
        provider: AI提供商实例
        role: 角色
        provider_name: 提供商名称
        selected_model: 选择的模型
        folder_path: 文档文件夹路径
    """
    print_info(f"使用 {provider_name} - {selected_model} 生成测试问题")
    print_info(f"文档路径: {folder_path}")
    console.print()

    # 读取文档
    file_names, file_contents = read_markdown_files(folder_path)

    if not file_names:
        print_error("没有找到可处理的文档")
        return

    # 生成输出文件名（使用时间戳）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"question_generation_{timestamp}.xlsx"

    # 存储所有生成的问题
    all_questions = []
    total_questions = 0
    total_docs = len(file_names)

    print_info(f"将在每处理完一个文档后自动保存到: {output_filename}")
    console.print()

    # 依次处理每个文档
    for idx, doc_name in enumerate(file_names, start=1):
        content = file_contents[doc_name]

        print_info(f"[{idx}/{total_docs}] 处理文档: {doc_name}")

        # 生成问题
        questions = generate_questions_for_document(
            provider=provider,
            model=selected_model,
            role=role,
            document_name=doc_name,
            document_content=content,
        )

        # 保存问题到列表
        if questions:
            for question in questions:
                all_questions.append((doc_name, question))
                total_questions += 1

            # 立即保存到Excel（增量保存，避免进度丢失）
            print_info(f"保存进度到 {output_filename}...")
            if export_questions_to_excel(all_questions, output_filename):
                print_success(f"✓ 已保存 {total_questions} 个问题 (来自 {idx} 个文档)")
            else:
                print_warning("保存失败，但会继续处理下一个文档")
        else:
            print_warning(f"文档 {doc_name} 未生成任何问题")

        console.print()


def run_cross_knowledge_generation(
    provider, role: str, provider_name: str, selected_model: str, folder_path: str
):
    """
    运行跨知识点问题生成流程

    Args:
        provider: AI提供商实例
        role: 角色
        provider_name: 提供商名称
        selected_model: 选择的模型
        folder_path: 文档文件夹路径
    """
    import random

    print_info(f"使用 {provider_name} - {selected_model} 生成跨知识点测试问题")
    print_info(f"文档路径: {folder_path}")
    console.print()

    # 读取文档
    file_names, file_contents = read_markdown_files(folder_path)

    if not file_names:
        print_error("没有找到可处理的文档")
        return

    # 预处理：将所有文档切分为 Chunks
    all_chunks = []  # List of (doc_name, chunk_content)
    for doc_name in file_names:
        content = file_contents[doc_name] or ""
        if not content:
            continue

        # 简单的按长度切分
        for i in range(0, len(content), MAX_DOC_CHARS_PER_CALL):
            chunk_text = content[i : i + MAX_DOC_CHARS_PER_CALL]
            all_chunks.append((doc_name, chunk_text))

    if len(all_chunks) < 1:
        print_error("文档内容过少，无法进行生成")
        return

    total_chunks = len(all_chunks)
    print_info(f"共加载 {len(file_names)} 个文档，切分为 {total_chunks} 个内容块")

    # 从配置中读取跨知识点生成迭代次数设置
    from dify_chat_tester.config_loader import get_config

    cfg = get_config()
    min_iterations = cfg.get_int("CROSS_KNOWLEDGE_MIN_ITERATIONS", 5)
    max_iterations = cfg.get_int("CROSS_KNOWLEDGE_MAX_ITERATIONS", 20)
    if min_iterations < 1:
        min_iterations = 1
    if max_iterations < min_iterations:
        max_iterations = min_iterations

    # 生成输出文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"cross_knowledge_questions_{timestamp}.xlsx"

    all_questions = []
    total_generated = 0

    # 设定迭代次数：
    # - 至少为配置中的 CROSS_KNOWLEDGE_MIN_ITERATIONS；
    # - 同时不超过 CROSS_KNOWLEDGE_MAX_ITERATIONS；
    # - 也会参考总 chunk 数量以保证一定覆盖度。
    iterations = max(total_chunks, min_iterations)
    if iterations > max_iterations:
        iterations = max_iterations

    print_info(f"计划执行 {iterations} 轮随机组合生成...")
    console.print()

    for i in range(iterations):
        # 随机抽取 2-3 个 chunk
        sample_size = random.randint(2, 3)
        # 如果 chunk 总数不够，就全取
        if total_chunks <= sample_size:
            selected_indices = range(total_chunks)
        else:
            selected_indices = random.sample(range(total_chunks), sample_size)

        selected_chunks = [all_chunks[idx] for idx in selected_indices]

        # 构建组合名称
        source_names = sorted(list(set([c[0] for c in selected_chunks])))
        doc_name_display = " + ".join(source_names)

        print_info(f"[轮次 {i+1}/{iterations}] 组合来源: {doc_name_display}")

        # 生成问题
        questions = generate_cross_doc_questions(
            provider=provider,
            model=selected_model,
            role=role,
            source_chunks=selected_chunks,
        )

        if questions:
            for q in questions:
                all_questions.append((f"跨知识点: {doc_name_display}", q))
                total_generated += 1

            print_success(f"  ✓ 生成 {len(questions)} 个问题")

            # 增量保存
            export_questions_to_excel(all_questions, output_filename)
        else:
            print_warning("  - 未生成有效问题 (可能内容相关性不足)")

        console.print()

    # 最终总结
    if all_questions:
        print_success("全部完成！")
        print_info(f"总计生成 {total_generated} 个跨知识点问题")
        print_info(f"结果已保存到: {output_filename}")
    else:
        print_warning("没有生成任何问题")


def generate_cross_doc_questions(
    provider, model: str, role: str, source_chunks: List[Tuple[str, str]]
) -> List[str]:
    """
    根据多个文本块生成跨知识点问题
    """
    # 构建上下文内容
    context_text = ""
    for idx, (doc_name, content) in enumerate(source_chunks, 1):
        context_text += f"\n--- 知识点来源 {idx} (文档: {doc_name}) ---\n{content}\n"

    # 从配置中获取提示词模板
    from dify_chat_tester.config_loader import get_config

    prompt_template = get_config().get_cross_knowledge_prompt()

    try:
        prompt = prompt_template.format(context_text=context_text)
    except KeyError as e:
        print_error(f"跨知识点提示词模板格式错误: {e}，将使用默认模板")
        prompt = f"""你是一个专业的测试问题生成助手。请阅读以下来自不同（或相同）文档的多个知识点片段，尝试寻找它们之间的关联，生成跨知识点的测试问题。

要求：
1. 分析提供的多个知识点来源，寻找逻辑关联（例如：操作流程的前后步骤、概念的对比、功能组合使用、不同角色的视角差异、同一功能在不同场景的表现等）。
2. 评估这些关联在用户真实场景中出现的概率。
   - 只有当概率为“中”或“高”时，才生成问题。
   - 如果片段之间完全不相关，或关联非常牵强/生僻，请不要生成问题。
3. 生成的问题必须需要结合多个知识点来源的内容才能完整回答（即跨知识点问题）。
4. 问题要模仿真实用户的自然提问。
5. 以 JSON 数组格式返回问题列表。如果没有合适的问题，返回空数组 []。

{context_text}

请生成问题列表（必须是 JSON 数组格式，例如: ["问题1", "问题2"]）："""

    try:
        response, success, error_msg, _ = provider.send_message(
            message=prompt,
            model=model,
            role=role,
            stream=False,
            show_indicator=True,
        )

        if not success:
            print_error(f"AI调用失败: {error_msg}")
            return []

        return parse_questions_from_response(response)

    except Exception as e:
        print_error(f"生成问题时发生错误: {e}")
        return []
