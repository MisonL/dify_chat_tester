"""Excel 工具模块
提供 Excel 相关的辅助函数，如安全写入单元格、初始化日志文件、清理非法字符等。
"""

import os

import openpyxl
from openpyxl.cell.cell import MergedCell


def write_cell_safely(worksheet, row, col, value):
    """
    安全地写入 Excel 单元格，处理合并单元格的情况。
    如果目标单元格是合并单元格的一部分，则写入合并区域的左上角单元格。
    """
    cell_obj = worksheet.cell(row=row, column=col)
    if isinstance(cell_obj, MergedCell):
        # 如果是合并单元格的一部分，找到其合并区域的左上角单元格
        for merged_range in worksheet.merged_cells.ranges:
            if cell_obj.coordinate in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                worksheet.cell(row=min_row, column=min_col).value = value
                return
    else:
        cell_obj.value = value


def init_excel_log(file_name, headers):
    """初始化 Excel 日志文件"""
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        if worksheet is None:
            worksheet = workbook.create_sheet("Log")

        # 设置表头
        worksheet.append(headers)

    return workbook, worksheet


def clean_excel_text(text):
    """清理文本中的 Excel 非法字符

    Excel 不允许以下控制字符：
    - 0x00, 0x01, ..., 0x08
    - 0x0B, 0x0C, 0x0D, 0x0E, ..., 0x1F
    - 0x7F
    """
    if text is None:
        return ""

    # 将文本转换为字符串（如果不是的话）
    text = str(text)

    # 移除非法字符
    # 保留 \t (0x09), \n (0x0A), \r (0x0D)
    illegal_chars = [
        "\x00",
        "\x01",
        "\x02",
        "\x03",
        "\x04",
        "\x05",
        "\x06",
        "\x07",
        "\x08",
        "\x0b",
        "\x0c",
        "\x0e",
        "\x0f",
        "\x10",
        "\x11",
        "\x12",
        "\x13",
        "\x14",
        "\x15",
        "\x16",
        "\x17",
        "\x18",
        "\x19",
        "\x1a",
        "\x1b",
        "\x1c",
        "\x1d",
        "\x1e",
        "\x1f",
        "\x7f",
    ]

    for char in illegal_chars:
        text = text.replace(char, "")

    return text


def log_to_excel(worksheet, row_data):
    """记录到 Excel（清理非法字符）"""
    # 清理每行数据中的非法字符
    cleaned_data = []
    for item in row_data:
        cleaned_data.append(clean_excel_text(item))

    worksheet.append(cleaned_data)
