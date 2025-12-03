"""
æ‰¹é‡æŸ¥è¯¢ç®¡ç†æ¨¡å—
è´Ÿè´£å¤„ç†æ‰¹é‡è¯¢é—®æ¨¡å¼çš„åŠŸèƒ½
"""

import json
import os
import time
from datetime import datetime

import openpyxl

from dify_chat_tester.config_loader import get_config
from dify_chat_tester.excel_utils import init_excel_log, log_to_excel
from dify_chat_tester.terminal_ui import (
    Panel,
    Text,
    box,
    console,
    print_error,
    print_file_list,
    print_input_prompt,
    print_statistics,
    print_success,
)

# æ¯å¤šå°‘æ¡é—®é¢˜ä¿å­˜ä¸€æ¬¡æ‰¹é‡æ—¥å¿—
SAVE_EVERY_N_QUERIES = 10


def run_batch_query(
    provider,
    selected_role: str,
    provider_name: str,
    selected_model: str,
    batch_request_interval: float,
    batch_default_show_response: bool,
):
    """è¿è¡Œæ‰¹é‡è¯¢é—®æ¨¡å¼"""
    # è·å–é…ç½®
    config = get_config()
    enable_thinking = config.get_enable_thinking()
    console.print()

    # æ¨¡å¼ä¿¡æ¯é¢æ¿
    mode_text = Text()
    mode_text.append("ğŸ¤– æ¨¡å‹: ", style="bold yellow")
    mode_text.append(f"{selected_model}\n", style="bold cyan")
    mode_text.append("ğŸ‘¤ è§’è‰²: ", style="bold yellow")
    mode_text.append(f"{selected_role}\n", style="bold cyan")
    mode_text.append("ğŸ’¬ ä¾›åº”å•†: ", style="bold yellow")
    mode_text.append(f"{provider_name}", style="bold cyan")

    # Dify ä¸å†éœ€è¦æ˜¾ç¤ºåº”ç”¨ ID

    mode_panel = Panel(
        mode_text,
        title="[bold]ğŸ“„ æ‰¹é‡è¯¢é—®æ¨¡å¼[/bold]",
        border_style="bright_magenta",
        box=box.ROUNDED,
        padding=(1, 2),
    )
    console.print(mode_panel)
    console.print()

    # åˆ—å‡ºå½“å‰ç›®å½•ä¸‹çš„ Excel æ–‡ä»¶
    excel_files = [
        f for f in os.listdir(".") if f.endswith(".xlsx") and os.path.isfile(f)
    ]

    selected_excel_file = None
    while True:
        if excel_files:
            print_file_list(excel_files)
            file_input = print_input_prompt("è¯·è¾“å…¥ Excel æ–‡ä»¶åºå·æˆ–ç›´æ¥è¾“å…¥æ–‡ä»¶è·¯å¾„")

            try:
                file_index = int(file_input)
                if 1 <= file_index <= len(excel_files):
                    excel_file_path = excel_files[file_index - 1]
                else:
                    print(
                        f"é”™è¯¯: æ— æ•ˆçš„æ–‡ä»¶åºå· '{file_index}'ã€‚è¯·é‡æ–°è¾“å…¥ã€‚",
                        file=console.file,
                    )
                    continue
            except ValueError:
                # ç”¨æˆ·è¾“å…¥çš„æ˜¯è·¯å¾„
                excel_file_path = file_input
        else:
            excel_file_path = print_input_prompt(
                "å½“å‰ç›®å½•ä¸‹æ²¡æœ‰æ‰¾åˆ° Excel æ–‡ä»¶ï¼Œè¯·è¾“å…¥åŒ…å«è¯¢é—®å†…å®¹çš„ Excel æ–‡ä»¶è·¯å¾„"
            )

        if not os.path.exists(excel_file_path):
            print(
                f"é”™è¯¯: æ–‡ä»¶ '{excel_file_path}' ä¸å­˜åœ¨ã€‚è¯·é‡æ–°è¾“å…¥ã€‚",
                file=console.file,
            )
            continue

        try:
            batch_workbook = openpyxl.load_workbook(excel_file_path)
            batch_worksheet = batch_workbook.active
            if batch_worksheet is None:  # ç¡®ä¿å·¥ä½œè¡¨ä¸ä¸ºNone
                print(
                    f"é”™è¯¯: Excel æ–‡ä»¶ '{excel_file_path}' ä¸­æ²¡æœ‰æ´»åŠ¨å·¥ä½œè¡¨ã€‚è¯·é‡æ–°è¾“å…¥ã€‚",
                    file=console.file,
                )
                continue
            selected_excel_file = excel_file_path
            break  # æˆåŠŸè¯»å–æ–‡ä»¶å¹¶è·å–å·¥ä½œè¡¨ï¼Œè·³å‡ºå¾ªç¯
        except Exception as e:
            print(
                f"é”™è¯¯: æ— æ³•è¯»å– Excel æ–‡ä»¶ '{excel_file_path}'ã€‚è¯·ç¡®ä¿æ–‡ä»¶æ ¼å¼æ­£ç¡®ä¸”æœªè¢«å ç”¨ã€‚é”™è¯¯ä¿¡æ¯: {e}ã€‚è¯·é‡æ–°è¾“å…¥ã€‚",
                file=console.file,
            )
            continue

    # ä¸ºå½“å‰è¾“å…¥æ–‡ä»¶æ„å»ºè¿›åº¦çŠ¶æ€æ–‡ä»¶è·¯å¾„
    state_dir = os.path.dirname(selected_excel_file) or "."
    state_file_name = f".batch_state_{os.path.basename(selected_excel_file)}.json"
    state_file_path = os.path.join(state_dir, state_file_name)

    # é»˜è®¤ä»ç¬¬äºŒè¡Œå¼€å§‹ï¼ˆç¬¬ä¸€è¡Œä¸ºè¡¨å¤´ï¼‰
    resume_from_row = 2
    batch_state = None

    if os.path.exists(state_file_path):
        try:
            with open(state_file_path, "r", encoding="utf-8") as f:
                loaded_state = json.load(f)
        except Exception:
            loaded_state = None

        if loaded_state and os.path.abspath(
            loaded_state.get("excel_file_path", "")
        ) == os.path.abspath(selected_excel_file):
            last_row = int(loaded_state.get("last_processed_row", 1))
            # ä»…å½“ä¹‹å‰ç¡®å®å¤„ç†è¿‡æ•°æ®è¡Œæ—¶æ‰æä¾›æ¢å¤é€‰é¡¹
            if last_row >= 2 and last_row < batch_worksheet.max_row:
                resume_choice = (
                    print_input_prompt(
                        f"æ£€æµ‹åˆ°å†å²è¿›åº¦è®°å½•ï¼šå·²å¤„ç†åˆ°ç¬¬ {last_row} è¡Œï¼Œæ˜¯å¦ä»ä¸‹ä¸€è¡Œç»§ç»­ï¼Ÿ(Y/n)"
                    )
                    .strip()
                    .lower()
                )
                if not resume_choice or resume_choice in ("y", "yes"):
                    resume_from_row = last_row + 1
                    batch_state = loaded_state
                else:
                    batch_state = None

    # è·å–åˆ—å
    column_names = [cell.value for cell in batch_worksheet[1]]
    print_success(f"å·²é€‰æ‹©æ–‡ä»¶: {selected_excel_file}")

    # æ£€æŸ¥æ˜¯å¦å­˜åœ¨â€œæ–‡æ¡£åç§°â€åˆ—ï¼ˆå¦‚æœå­˜åœ¨åˆ™å¤ç”¨ï¼‰
    doc_name_col_index = None
    try:
        if batch_state and "doc_name_col_index" in batch_state:
            doc_name_col_index = int(batch_state["doc_name_col_index"])
        else:
            doc_name_col_index = column_names.index("æ–‡æ¡£åç§°")
    except ValueError:
        doc_name_col_index = None

    # è®©ç”¨æˆ·é€šè¿‡åºå·é€‰æ‹©é—®é¢˜åˆ—
    from dify_chat_tester.terminal_ui import select_column_by_index

    if batch_state and "question_col_index" in batch_state:
        question_col_index = int(batch_state["question_col_index"])
    else:
        question_col_index = select_column_by_index(
            column_names, "è¯·é€‰æ‹©é—®é¢˜æ‰€åœ¨åˆ—çš„åºå·"
        )

    # æ³¨æ„ï¼šä¸å†åˆ›å»ºæˆ–ä½¿ç”¨å›ç­”åˆ—ï¼Œæ‰€æœ‰ç»“æœåªè®°å½•åˆ°æ—¥å¿—æ–‡ä»¶

    # è¯¢é—®æ˜¯å¦æ˜¾ç¤ºæ¯ä¸ªé—®é¢˜çš„å›ç­”å†…å®¹ï¼ˆå›è½¦åˆ™ä½¿ç”¨é…ç½®ä¸­çš„é»˜è®¤å€¼ï¼‰
    display_response_choice = print_input_prompt(
        "æ˜¯å¦åœ¨æ§åˆ¶å°æ˜¾ç¤ºæ¯ä¸ªé—®é¢˜çš„å›ç­”å†…å®¹ï¼Ÿ (Y/nï¼Œç›´æ¥å›è½¦ä½¿ç”¨é…ç½®é»˜è®¤å€¼)"
    )
    if not display_response_choice:
        show_batch_response = batch_default_show_response
    else:
        show_batch_response = display_response_choice.lower() != "n"

    # ä»é…ç½®ä¸­è·å–è¯·æ±‚é—´éš”æ—¶é—´ï¼ˆä½¿ç”¨é…ç½®ä¸­çš„é»˜è®¤å€¼ï¼‰
    request_interval = batch_request_interval

    # ä¸ºæ‰¹é‡æ¨¡å¼è®¾ç½® show_indicatorï¼ˆåªæœ‰åœ¨ç”¨æˆ·é€‰æ‹©æ˜¾ç¤ºå“åº”æ—¶æ‰å¯ç”¨ï¼‰
    batch_show_indicator = show_batch_response

    # è¯¦ç»†æ—¥å¿—æ–‡ä»¶ï¼Œç”¨äºè®°å½•æ¯æ¬¡è¯·æ±‚çš„è¯¦ç»†ä¿¡æ¯
    if (
        batch_state
        and batch_state.get("batch_log_file")
        and os.path.exists(batch_state["batch_log_file"])
    ):
        output_file_name = batch_state["batch_log_file"]
    else:
        output_file_name = (
            f"batch_query_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    batch_log_headers = [
        "æ—¶é—´æˆ³",
        "è§’è‰²",
        "æ–‡æ¡£åç§°",
        "åŸå§‹é—®é¢˜",
        f"{provider_name}å“åº”",
        "æ˜¯å¦æˆåŠŸ",
        "é”™è¯¯ä¿¡æ¯",
        "å¯¹è¯ID",
    ]
    output_workbook, output_worksheet = init_excel_log(
        output_file_name, batch_log_headers
    )

    total_queries = 0
    successful_queries = 0
    failed_queries = 0
    # è‡ªä¸Šæ¬¡ä¿å­˜ä»¥æ¥å·²å¤„ç†çš„é—®é¢˜æ•°é‡
    queries_since_last_save = 0
    start_time = time.time()

    total_rows = batch_worksheet.max_row - 1

    # å¦‚æœä¸Šæ¬¡å·²ç»å¤„ç†åˆ°æ–‡ä»¶æœ«å°¾ï¼Œåˆ™ç›´æ¥ç»“æŸ
    if resume_from_row > batch_worksheet.max_row:
        print_success("æ£€æµ‹åˆ°è¯¥æ–‡ä»¶çš„æ‰€æœ‰é—®é¢˜å‡å·²å¤„ç†å®Œæˆï¼Œæ— éœ€ç»§ç»­ã€‚")
        # æ¸…ç†å†å²è¿›åº¦æ–‡ä»¶
        try:
            if os.path.exists(state_file_path):
                os.remove(state_file_path)
        except Exception:
            pass
        return

    print(
        f"\nå¼€å§‹æ‰¹é‡è¯¢é—®... (å…± {total_rows} è¡Œæ•°æ®ï¼Œå½“å‰ä»ç¬¬ {resume_from_row} è¡Œå¼€å§‹)"
    )
    for row_idx in range(
        resume_from_row, batch_worksheet.max_row + 1
    ):  # ä»æŒ‡å®šè¡Œå¼€å§‹è¯»å–æ•°æ®
        # è·å–æ–‡æ¡£åç§°ï¼ˆå¦‚æœè¾“å…¥è¡¨ä¸­å­˜åœ¨å¯¹åº”åˆ—ï¼‰
        doc_name = ""
        if doc_name_col_index is not None:
            doc_cell_value = batch_worksheet.cell(
                row=row_idx, column=doc_name_col_index + 1
            ).value
            doc_name = str(doc_cell_value) if doc_cell_value is not None else ""

        question_cell_value = batch_worksheet.cell(
            row=row_idx, column=question_col_index + 1
        ).value
        question = (
            str(question_cell_value) if question_cell_value is not None else ""
        )  # ç¡®ä¿è½¬æ¢ä¸ºå­—ç¬¦ä¸²

        if not question.strip():  # æ£€æŸ¥é—®é¢˜æ˜¯å¦ä¸ºç©ºæˆ–åªåŒ…å«ç©ºæ ¼
            print(f"è­¦å‘Š: ç¬¬ {row_idx} è¡Œé—®é¢˜ä¸ºç©ºï¼Œè·³è¿‡ã€‚", file=console.file)
            failed_queries += 1  # ç©ºé—®é¢˜ä¹Ÿç®—ä½œå¤±è´¥
            log_to_excel(
                output_worksheet,
                [
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    selected_role,
                    doc_name,
                    question,  # åŸå§‹é—®é¢˜ä¸ºç©º
                    "",
                    False,
                    "é—®é¢˜ä¸ºç©º",
                    0,
                    "",
                ],
            )
            continue  # è·³è¿‡å½“å‰å¾ªç¯çš„å‰©ä½™éƒ¨åˆ†

        total_queries += 1  # åªæœ‰éç©ºé—®é¢˜æ‰è®¡å…¥æ€»æ•°

        # è®¡ç®—è¿›åº¦
        current_progress = row_idx - 1
        pending_count = total_rows - current_progress
        progress_percent = (current_progress / total_rows) * 100

        # ç¾åŒ–é—®é¢˜æ˜¾ç¤ºï¼ˆåŠ ç²—å’Œé¢œè‰²ï¼‰
        question_display = (
            f"[bold bright_magenta]å¤„ç†è¿›åº¦ ({current_progress}/{total_rows} - {progress_percent:.1f}%) "
            f"| å¾…å¤„ç†: {pending_count} | é—®é¢˜:[/bold bright_magenta] "
            f"[bold yellow]{question[:50]}{'...' if len(question) > 50 else ''}[/bold yellow]"
        )
        console.print(f"\n{question_display}")

        response, success, error, conversation_id = provider.send_message(
            message=question,
            model=selected_model,
            role=selected_role,
            stream=True,
            show_indicator=batch_show_indicator,
            show_thinking=enable_thinking,
        )

        if success:
            successful_queries += 1
            print(f"é—®é¢˜ (ç¬¬ {total_queries} ä¸ª) å¤„ç†å®Œæˆã€‚")  # ç®€æ´æç¤º
        else:
            failed_queries += 1
            print(f"é—®é¢˜ (ç¬¬ {total_queries} ä¸ª) å¤„ç†å¤±è´¥ã€‚é”™è¯¯: {error}")  # ç®€æ´æç¤º

        # è®°å½•è¯¦ç»†æ—¥å¿—åˆ°æ—¥å¿—æ–‡ä»¶
        log_to_excel(
            output_worksheet,
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                selected_role,
                doc_name,
                question,
                response,
                success,
                error,
                conversation_id or "",
            ],
        )

        # æŒ‰æ‰¹æ¬¡ä¿å­˜æ—¥å¿—ï¼Œå‡å°‘ç£ç›˜ IO
        queries_since_last_save += 1
        if queries_since_last_save >= SAVE_EVERY_N_QUERIES:
            try:
                output_workbook.save(output_file_name)
                queries_since_last_save = 0
            except PermissionError:
                print_error(
                    f"è­¦å‘Šï¼šæ— æ³•ä¿å­˜æ—¥å¿—æ–‡ä»¶ '{output_file_name}'ã€‚è¯·ç¡®ä¿æ–‡ä»¶æœªè¢«å…¶ä»–ç¨‹åºæ‰“å¼€ã€‚"
                )
            except Exception as e:
                print_error(f"è­¦å‘Šï¼šä¿å­˜æ—¥å¿—æ—¶å‡ºé”™ï¼š{e}")

        # æ¯å¤„ç†ä¸€è¡Œæ›´æ–°ä¸€æ¬¡è¿›åº¦çŠ¶æ€ï¼Œä¾¿äºä¸­æ–­åæ¢å¤
        try:
            current_state = {
                "excel_file_path": os.path.abspath(selected_excel_file),
                "batch_log_file": output_file_name,
                "last_processed_row": row_idx,
                "question_col_index": question_col_index,
                "doc_name_col_index": doc_name_col_index,
                "total_rows": batch_worksheet.max_row,
                "provider_name": provider_name,
                "selected_model": selected_model,
                "selected_role": selected_role,
            }
            with open(state_file_path, "w", encoding="utf-8") as f:
                json.dump(current_state, f, ensure_ascii=False, indent=2)
        except Exception:
            # è¿›åº¦æ–‡ä»¶å†™å…¥å¤±è´¥ä¸å½±å“ä¸»æµç¨‹
            pass

        # åŸå§‹æ–‡ä»¶ä¿æŒä¸å˜ï¼Œåªè®°å½•åˆ°æ—¥å¿—æ–‡ä»¶

        time.sleep(request_interval)  # é—´éš”æ—¶é—´

    # å¾ªç¯ç»“æŸååšä¸€æ¬¡æœ€ç»ˆä¿å­˜
    try:
        output_workbook.save(output_file_name)
    except PermissionError:
        print_error(
            f"è­¦å‘Šï¼šæ— æ³•ä¿å­˜æ—¥å¿—æ–‡ä»¶ '{output_file_name}'ã€‚è¯·ç¡®ä¿æ–‡ä»¶æœªè¢«å…¶ä»–ç¨‹åºæ‰“å¼€ã€‚"
        )
    except Exception as e:
        print_error(f"è­¦å‘Šï¼šä¿å­˜æ—¥å¿—æ—¶å‡ºé”™ï¼š{e}")

    # æ‰¹é‡æ‰§è¡Œå…¨éƒ¨å®Œæˆåï¼Œæ¸…ç†è¿›åº¦çŠ¶æ€æ–‡ä»¶
    try:
        if os.path.exists(state_file_path):
            os.remove(state_file_path)
    except Exception:
        pass

    end_time = time.time()
    total_duration = end_time - start_time

    # ç»Ÿè®¡ä¿¡æ¯é¢æ¿
    print_statistics(total_queries, successful_queries, failed_queries, total_duration)

    # æ‰§è¡Œä¿¡æ¯æ±‡æ€»é¢æ¿
    summary_text = Text()
    summary_text.append("ğŸ“ æ–‡ä»¶ä¿¡æ¯\n", style="bold yellow")
    summary_text.append(f"  â€¢ å¤„ç†æ–‡ä»¶: {selected_excel_file}\n", style="white")
    summary_text.append(
        f"  â€¢ é—®é¢˜åˆ—: {column_names[question_col_index]} (ç¬¬{question_col_index + 1}åˆ—)\n",
        style="white",
    )
    summary_text.append("  â€¢ æ—¥å¿—æ–‡ä»¶: æ‰€æœ‰ç»“æœä¿å­˜åˆ°ç‹¬ç«‹æ—¥å¿—æ–‡ä»¶\n\n", style="white")

    summary_text.append("ğŸ¤– æ¨¡å‹é…ç½®\n", style="bold yellow")
    summary_text.append(f"  â€¢ AI ä¾›åº”å•†: {provider_name}\n", style="white")
    summary_text.append(f"  â€¢ é€‰ç”¨æ¨¡å‹: {selected_model}\n", style="white")
    summary_text.append(f"  â€¢ è§’è‰²è®¾å®š: {selected_role}\n", style="white")

    # æ·»åŠ  API æ¥å£åœ°å€
    base_url = getattr(provider, "base_url", None)
    if base_url:
        summary_text.append(f"  â€¢ API æ¥å£: {base_url}\n", style="white")

    # Dify ä¸å†éœ€è¦æ˜¾ç¤ºåº”ç”¨ ID

    summary_text.append("\n", style="white")

    success_rate = (
        (successful_queries / total_queries * 100) if total_queries > 0 else 0
    )
    summary_text.append("ğŸ“Š æ‰§è¡Œç»Ÿè®¡\n", style="bold yellow")
    summary_text.append(f"  â€¢ æˆåŠŸç‡: {success_rate:.1f}%\n", style="white")
    summary_text.append(f"  â€¢ è¯·æ±‚é—´éš”: {batch_request_interval}ç§’\n", style="white")
    summary_text.append(f"  â€¢ è¯¦ç»†æ—¥å¿—: {output_file_name}", style="white")

    summary_panel = Panel(
        summary_text,
        title="[bold]ğŸ“‹ æ‰§è¡Œä¿¡æ¯æ±‡æ€»[/bold]",
        border_style="bright_magenta",
        box=box.ROUNDED,
        padding=(1, 2),
    )
    console.print()
    console.print(summary_panel)
    console.print()
