"""
AI 聊天客户端测试工具（支持多AI供应商）

作者：Mison
邮箱：1360962086@qq.com
仓库：https://github.com/MisonL/dify_chat_tester
许可证：MIT
"""

import openpyxl
from openpyxl.cell.cell import MergedCell # 导入 MergedCell 类型
from datetime import datetime
import os
import sys
import time
import re
import urllib.parse

from dify_chat_tester.ai_providers import get_provider, AIProvider
from dify_chat_tester.config_loader import get_config
from dify_chat_tester.terminal_ui import (
    print_success, print_error, print_warning, print_info,
    print_input_prompt, create_provider_menu,
    print_statistics, print_welcome, print_api_key_confirmation, print_file_list,
    select_column_by_index, Icons, console, Text, Panel, box, input_api_key
)
from dify_chat_tester import __version__
from rich.prompt import Confirm
import requests


def hide_api_key(key: str) -> str:
    """
    隐藏 API 密钥的中间部分，只显示前4位和后4位
    例如: sk-1234567890abcdef 变成 sk-1234****89abcdef

    Args:
        key: 原始 API 密钥

    Returns:
        隐藏后的密钥字符串
    """
    if len(key) <= 8:
        # 密钥太短，全部隐藏
        return "*" * len(key)

    # 显示前4位和后4位
    return key[:4] + "*" * (len(key) - 8) + key[-4:]


# input_api_key 函数已移动到 terminal_ui.py 模块中

# 显示版本信息
version_text = Text(f"dify_chat_tester v{__version__}", style="bold cyan")
version_panel = Panel(
    version_text,
    box=box.ROUNDED,
    padding=(0, 1),
    border_style="cyan",
    width=78  # 与艺术字宽度匹配
)
console.print(version_panel)

# 显示项目信息
info_text = Text()
info_text.append("作者: ", style="bold")
info_text.append("Mison", style="white")
info_text.append("  | ", style="dim")
info_text.append("许可证: ", style="bold")
info_text.append("MIT", style="green")
info_text.append("  | ", style="dim")
info_text.append("邮箱: ", style="bold")
info_text.append("1360962086@qq.com", style="blue")
info_text.append("\n", style="dim")
info_text.append("项目地址: ", style="bold")
info_text.append("https://github.com/MisonL/dify_chat_tester", style="blue underline")

info_panel = Panel(
    info_text,
    box=box.ROUNDED,
    padding=(0, 1),
    border_style="dim",
    width=78  # 与艺术字宽度匹配
)
console.print(info_panel)
console.print()

# 加载配置
config = get_config()

# ========================================
# 📋 配置加载完成
# ========================================
# 所有配置现在从 config.env 文件中加载
# 修改 config.env 文件后重启程序生效
# 详细说明请查看 "配置说明.md" 文件

# 从配置中获取值
CHAT_LOG_FILE_NAME = config.get_str('CHAT_LOG_FILE_NAME', 'chat_log.xlsx')
ROLES = config.get_list('ROLES', default=['员工', '门店'])

# 特殊解析 AI_PROVIDERS 配置（格式：序号:名称:ID;序号:名称:ID）
def parse_ai_providers(value: str) -> dict:
    """解析 AI_PROVIDERS 配置（格式：序号:名称:ID）"""
    default = {
        "1": {"name": "Dify", "id": "dify"},
        "2": {"name": "OpenAI 兼容接口", "id": "openai"},
        "3": {"name": "iFlow", "id": "iflow"}
    }
    if not value:
        return default

    result = {}
    for item in value.split(';'):
        if ':' in item:
            parts = item.split(':', 2)  # 最多分割2次，得到3个部分
            if len(parts) == 3:
                key, name, provider_id = parts
                result[key.strip()] = {
                    "name": name.strip(),
                    "id": provider_id.strip()
                }
    return result if result else default

AI_PROVIDERS = parse_ai_providers(config.get_str('AI_PROVIDERS', ''))

# 批量询问配置
BATCH_REQUEST_INTERVAL = config.get_float('BATCH_REQUEST_INTERVAL', 1.0)
BATCH_DEFAULT_SHOW_RESPONSE = config.get_bool('BATCH_DEFAULT_SHOW_RESPONSE', False)

def write_cell_safely(worksheet, row, col, value):
    """
    安全地写入 Excel 单元格，处理合并单元格的情况。
    如果目标单元格是合并单元格的一部分，则写入合并区域的左上角单元格。
    """
    cell_obj = worksheet.cell(row=row, column=col)
    if isinstance(cell_obj, MergedCell):
        # 如果是合并单元格的一部分，找到其合并区域的左上角单元格
        for merged_range in worksheet.merged_cells.ranges:
            if cell_obj.coordinate in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                worksheet.cell(row=min_row, column=min_col).value = value
                return
    else:
        cell_obj.value = value

def init_excel_log(file_name, headers):
    """初始化 Excel 日志文件"""
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        if worksheet is None:
            worksheet = workbook.create_sheet("Log")
        
        # 设置表头
        worksheet.append(headers)
    
    return workbook, worksheet

def clean_excel_text(text):
    """清理文本中的 Excel 非法字符
    
    Excel 不允许以下控制字符：
    - 0x00, 0x01, ..., 0x08
    - 0x0B, 0x0C, 0x0D, 0x0E, ..., 0x1F
    - 0x7F
    """
    if text is None:
        return ""
    
    # 将文本转换为字符串（如果不是的话）
    text = str(text)
    
    # 移除非法字符
    # 保留 \t (0x09), \n (0x0A), \r (0x0D)
    illegal_chars = [
        '\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', '\x08',
        '\x0B', '\x0C', '\x0E', '\x0F', '\x10', '\x11', '\x12', '\x13', '\x14',
        '\x15', '\x16', '\x17', '\x18', '\x19', '\x1A', '\x1B', '\x1C', '\x1D',
        '\x1E', '\x1F', '\x7F'
    ]
    
    for char in illegal_chars:
        text = text.replace(char, '')
    
    return text

def log_to_excel(worksheet, row_data):
    """记录到 Excel（清理非法字符）"""
    # 清理每行数据中的非法字符
    cleaned_data = []
    for item in row_data:
        cleaned_data.append(clean_excel_text(item))
    
    worksheet.append(cleaned_data)

def run_interactive_chat(provider: AIProvider, selected_role: str, provider_name: str, selected_model: str):
    """运行会话模式"""
    # 初始化 Excel
    chat_headers = ["时间戳", "角色", "用户输入", f"{provider_name}响应", "是否成功", "错误信息", "对话轮次", "对话ID"]
    workbook, worksheet = init_excel_log(CHAT_LOG_FILE_NAME, chat_headers)

    print_success(f"已选择角色: {selected_role}")
    print_success(f"已选择模型: {selected_model}")
    console.print()
    console.print(f"{Icons.INFO} 命令说明:", style="bold cyan")
    console.print(f"  {Icons.USER} 输入 '/exit' 或 '/quit' 返回主菜单", style="white")
    console.print(f"  {Icons.USER} 输入 '/new' 开启新的对话（重置上下文）", style="white")
    console.print()

    # 多轮对话支持
    conversation_id = None  # 对话ID，用于维护多轮对话上下文
    conversation_round = 0  # 对话轮次计数器

    # 聊天循环
    while True:
        user_input = print_input_prompt(f"{Icons.USER} 你")
        user_input = user_input.strip()

        # 处理退出命令 - 返回主菜单
        if user_input.lower() in ["/exit", "/quit"]:
            console.print()
            print_info("正在返回主菜单...")
            # 关闭工作簿（日志已实时保存）
            try:
                workbook.close()
                if conversation_round > 0:  # 只有当有对话内容时才显示消息
                    print_success(f"对话已保存到 {CHAT_LOG_FILE_NAME} (共 {conversation_round} 轮对话)")
            except Exception as e:
                print_error(f"关闭日志文件时出错：{e}")
            return  # 返回到主菜单，而不是退出程序

        # 处理开启新对话命令
        if user_input == "/new":
            conversation_id = None
            conversation_round = 0
            console.print()
            print_success("已开启新对话（上下文已重置）")
            console.print()
            continue

        conversation_round += 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 发送到 AI 供应商
        response, success, error, new_conversation_id = provider.send_message(
            message=user_input,
            model=selected_model,
            role=selected_role,
            conversation_id=conversation_id,
            stream=True,
            show_indicator=True
        )

        # 更新对话ID（用于后续多轮对话）
        if new_conversation_id:
            conversation_id = new_conversation_id

        # 记录到 Excel
        log_to_excel(
            worksheet,
            [
                timestamp,
                selected_role,
                user_input,
                response,
                success,
                error,
                conversation_round,
                conversation_id or ""  # 确保传递字符串（None时用空字符串）
            ]
        )

        # 实时保存日志（每轮对话后都保存）
        try:
            workbook.save(CHAT_LOG_FILE_NAME)
        except PermissionError:
            print_error(f"警告：无法实时保存日志文件 '{CHAT_LOG_FILE_NAME}'。请确保文件未被其他程序打开。")
        except Exception as e:
            print_error(f"警告：保存日志时出错：{e}")

    # 注意：日志保存已在退出时处理

def run_batch_query(provider: AIProvider, selected_role: str, provider_name: str, selected_model: str):
    """运行批量询问模式"""
    console.print()
    
    # 模式信息面板
    mode_text = Text()
    mode_text.append("🤖 模型: ", style="bold yellow")
    mode_text.append(f"{selected_model}\n", style="bold cyan")
    mode_text.append("👤 角色: ", style="bold yellow")
    mode_text.append(f"{selected_role}\n", style="bold cyan")
    mode_text.append("💬 供应商: ", style="bold yellow")
    mode_text.append(f"{provider_name}", style="bold cyan")
    
    # 如果是 Dify，添加应用 ID
    app_id = getattr(provider, 'app_id', None)
    if provider_name == "Dify" and app_id:
        mode_text.append("\n🔑 应用 ID: ", style="bold yellow")
        mode_text.append(f"{app_id}", style="bold cyan")
    
    mode_panel = Panel(
        mode_text,
        title="[bold]📄 批量询问模式[/bold]",
        border_style="bright_magenta",
        box=box.ROUNDED,
        padding=(1, 2)
    )
    console.print(mode_panel)
    console.print()

    # 列出当前目录下的 Excel 文件
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and os.path.isfile(f)]

    selected_excel_file = None
    while True:
        if excel_files:
            print_file_list(excel_files)
            file_input = print_input_prompt("请输入 Excel 文件序号或直接输入文件路径")

            try:
                file_index = int(file_input)
                if 1 <= file_index <= len(excel_files):
                    excel_file_path = excel_files[file_index - 1]
                else:
                    print(f"错误: 无效的文件序号 '{file_index}'。请重新输入。", file=sys.stderr)
                    continue
            except ValueError:
                # 用户输入的是路径
                excel_file_path = file_input
        else:
            excel_file_path = print_input_prompt("当前目录下没有找到 Excel 文件，请输入包含询问内容的 Excel 文件路径")

        if not os.path.exists(excel_file_path):
            print(f"错误: 文件 '{excel_file_path}' 不存在。请重新输入。", file=sys.stderr)
            continue

        try:
            batch_workbook = openpyxl.load_workbook(excel_file_path)
            batch_worksheet = batch_workbook.active
            if batch_worksheet is None:  # 确保工作表不为None
                print(f"错误: Excel 文件 '{excel_file_path}' 中没有活动工作表。请重新输入。", file=sys.stderr)
                continue
            selected_excel_file = excel_file_path
            break  # 成功读取文件并获取工作表，跳出循环
        except Exception as e:
            print(f"错误: 无法读取 Excel 文件 '{excel_file_path}'。请确保文件格式正确且未被占用。错误信息: {e}。请重新输入。", file=sys.stderr)
            continue

    # 获取列名
    column_names = [cell.value for cell in batch_worksheet[1]]
    print_success(f"已选择文件: {selected_excel_file}")

    # 让用户通过序号选择问题列
    question_col_index = select_column_by_index(column_names, "请选择问题所在列的序号")

    # 注意：不再创建或使用回答列，所有结果只记录到日志文件

    # 询问是否显示每个问题的回答内容
    display_response_choice = print_input_prompt("是否在控制台显示每个问题的回答内容？ (y/N)")
    show_batch_response = (display_response_choice.lower() == 'y') if display_response_choice else BATCH_DEFAULT_SHOW_RESPONSE

    # 从配置中获取请求间隔时间（使用配置中的默认值）
    request_interval = BATCH_REQUEST_INTERVAL

    # 为批量模式设置 show_indicator（只有在用户选择显示响应时才启用）
    batch_show_indicator = show_batch_response

    # 详细日志文件，用于记录每次请求的详细信息
    output_file_name = f"batch_query_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    batch_log_headers = ["时间戳", "角色", "原始问题", f"{provider_name}响应", "是否成功", "错误信息", "对话ID"]
    output_workbook, output_worksheet = init_excel_log(output_file_name, batch_log_headers)

    total_queries = 0
    successful_queries = 0
    failed_queries = 0
    start_time = time.time()

    print("\n开始批量询问...")
    for row_idx in range(2, batch_worksheet.max_row + 1):  # 从第二行开始读取数据
        question_cell_value = batch_worksheet.cell(row=row_idx, column=question_col_index + 1).value
        question = str(question_cell_value) if question_cell_value is not None else ""  # 确保转换为字符串

        if not question.strip():  # 检查问题是否为空或只包含空格
            print(f"警告: 第 {row_idx} 行问题为空，跳过。", file=sys.stderr)
            failed_queries += 1  # 空问题也算作失败
            log_to_excel(output_worksheet, [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                selected_role,
                question,  # 原始问题为空
                "",
                False,
                "问题为空",
                0,
                ""
            ])
            continue  # 跳过当前循环的剩余部分

        total_queries += 1  # 只有非空问题才计入总数
        
        # 美化问题显示（加粗和颜色）
        question_display = f"[bold bright_magenta]处理问题 (第 {total_queries} 个):[/bold bright_magenta] [bold yellow]{question[:50]}{'...' if len(question) > 50 else ''}[/bold yellow]"
        console.print(f"\n{question_display}")

        response, success, error, conversation_id = provider.send_message(
            message=question,
            model=selected_model,
            role=selected_role,
            stream=True,
            show_indicator=batch_show_indicator
        )

        if success:
            successful_queries += 1
            print(f"问题 (第 {total_queries} 个) 处理完成。")  # 简洁提示
        else:
            failed_queries += 1
            print(f"问题 (第 {total_queries} 个) 处理失败。错误: {error}")  # 简洁提示

        # 记录详细日志到新的Excel文件
        log_to_excel(output_worksheet, [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            selected_role,
            question,
            response,
            success,
            error,
            1,  # 批量询问通常是单轮对话，这里设为1
            conversation_id or ""
        ])

        # 原始文件保持不变，只记录到日志文件

        time.sleep(request_interval)  # 间隔时间

    end_time = time.time()
    total_duration = end_time - start_time

    # 保存详细日志文件
    output_workbook.save(output_file_name)
    print_success(f"详细日志已保存到 {output_file_name}")
    
    # 统计信息面板
    print_statistics(total_queries, successful_queries, failed_queries, total_duration)
    
    # 执行信息汇总面板
    summary_text = Text()
    summary_text.append("📁 文件信息\n", style="bold yellow")
    summary_text.append(f"  • 处理文件: {selected_excel_file}\n", style="white")
    summary_text.append(f"  • 问题列: {column_names[question_col_index]} (第{question_col_index+1}列)\n", style="white")
    summary_text.append("  • 日志文件: 所有结果保存到独立日志文件\n\n", style="white")
    
    summary_text.append("🤖 模型配置\n", style="bold yellow")
    summary_text.append(f"  • AI 供应商: {provider_name}\n", style="white")
    summary_text.append(f"  • 选用模型: {selected_model}\n", style="white")
    summary_text.append(f"  • 角色设定: {selected_role}\n", style="white")
    
    # 添加 API 接口地址
    base_url = getattr(provider, 'base_url', None)
    if base_url:
        summary_text.append(f"  • API 接口: {base_url}\n", style="white")
    
    # 如果是 Dify，添加应用 ID
    app_id = getattr(provider, 'app_id', None)
    if provider_name == "Dify" and app_id:
        summary_text.append(f"  • 应用 ID: {app_id}\n", style="white")
    
    summary_text.append("\n", style="white")
    
    success_rate = (successful_queries / total_queries * 100) if total_queries > 0 else 0
    summary_text.append("📊 执行统计\n", style="bold yellow")
    summary_text.append(f"  • 成功率: {success_rate:.1f}%\n", style="white")
    summary_text.append(f"  • 请求间隔: {BATCH_REQUEST_INTERVAL}秒\n", style="white")
    summary_text.append(f"  • 详细日志: {output_file_name}", style="white")
    
    summary_panel = Panel(
        summary_text,
        title="[bold]📋 执行信息汇总[/bold]",
        border_style="bright_magenta",
        box=box.ROUNDED,
        padding=(1, 2)
    )
    console.print()
    console.print(summary_panel)
    console.print()

def main():
    """主函数"""
    # 打印欢迎信息
    print_welcome()

    # 选择 AI 供应商
    provider_choice = create_provider_menu(AI_PROVIDERS)
    provider_id = AI_PROVIDERS[provider_choice]["id"]
    provider_name = AI_PROVIDERS[provider_choice]["name"]
    print_success(f"已选择: {provider_name}")

    # 根据供应商获取配置
    provider = None

    if provider_id == "dify":
        # Dify 配置
        base_url = ""
        app_id = ""
        api_key = ""
        
        # 配置 API 地址
        print_info("请输入 Dify 服务器地址配置")
        while True:
            dify_info = Text()
            dify_info.append("• Dify Cloud: ", style="bold cyan")
            dify_info.append("https://api.dify.ai\n", style="white")
            dify_info.append("• 私有化部署: ", style="bold cyan")
            dify_info.append("输入API接口地址（如: http://your-domain.com/v1）\n", style="white")
            dify_info.append("• 获取方式: Dify工作流 > 左侧菜单 > 访问API > 右上角API服务器处复制", style="bold cyan")

            panel = Panel(dify_info, title=f"{Icons.INFO} Dify 服务器配置", box=box.ROUNDED)
            console.print(panel)

            dify_url_input = print_input_prompt("请输入Dify服务器API接口地址(可在dify工作流>左侧,访问API>右上角,API服务器处复制)").strip()

            if not dify_url_input:
                dify_url_input = "https://api.dify.ai/v1"
                print_success(f"使用默认值: {dify_url_input}")

            # 补全协议
            if not dify_url_input.startswith(('http://', 'https://')):
                dify_url_input = "https://" + dify_url_input

            parsed_url = urllib.parse.urlparse(dify_url_input)

            if not parsed_url.netloc:
                print_error("URL 格式无效。请确保输入有效的域名或 IP。")
                continue

            # 直接使用用户输入的完整 API 地址，不做任何修改
            base_url = dify_url_input
            
            # 验证 API 地址是否有效
            print_info("正在验证 API 地址...")
            try:
                # 发送一个简单的 GET 请求到 API 地址
                test_url = base_url.rstrip('/') + '/health'  # 尝试访问健康检查端点
                response = requests.get(test_url, timeout=5)
                
                # 如果健康检查成功
                if response.status_code == 200:
                    print_success(f"API 地址验证成功: {base_url}")
                else:
                    # 尝试访问基础路径
                    response = requests.get(base_url, timeout=5)
                    if response.status_code in [200, 404, 405]:  # 404/405 也说明地址可达
                        print_success(f"API 地址可达: {base_url}")
                    else:
                        print_warning(f"API 地址可能无效 (状态码: {response.status_code})")
                        if not Confirm.ask("是否继续使用此地址？", default=False):
                            continue
            except requests.exceptions.RequestException as e:
                print_error(f"API 地址验证失败: {str(e)}")
                print_info("请检查:")
                print_info("  • 地址是否正确")
                print_info("  • 网络是否可达")
                print_info("  • Dify 服务是否正在运行")
                if not Confirm.ask("是否忽略验证错误并继续？", default=False):
                    continue
            
            app_id_match = re.search(r'/app/([0-9a-fA-F-]{36})', parsed_url.path)
            if app_id_match:
                app_id = app_id_match.group(1)
                print_success(f"已从 URL 中提取 Dify 应用 ID: {app_id}")
            break  # API 地址配置完成
        
        # 配置应用 ID（如果未从 URL 中提取）
        if not app_id:
            while True:
                app_id = print_input_prompt("请输入 Dify 应用 ID")
                if not app_id:
                    print_error("Dify 应用 ID 不能为空。")
                    continue
                if not re.match(r"^[0-9a-fA-F-]{36}$", app_id):
                    print_warning("Dify 应用 ID 格式可能不正确（非标准 UUID 格式）。请确认。")
                    if not Confirm.ask("是否继续使用此应用 ID？", default=False):
                        continue
                break
        
        # 配置 API 密钥
        while True:
            print_info("请输入 Dify API 密钥（输入不会显示在屏幕上）")
            api_key = input_api_key("密钥")
            if not api_key:
                print_error("Dify API 密钥不能为空。")
                continue
            if not api_key.startswith("app-"):
                print_error("Dify API 密钥必须以 'app-' 开头。")
                continue
            # 确认输入的密钥（显示部分）
            if print_api_key_confirmation(hide_api_key(api_key)):
                break

        provider = get_provider("dify", base_url=base_url, api_key=api_key, app_id=app_id)

    elif provider_id == "openai":
        # OpenAI 兼容接口配置
        base_url = print_input_prompt("请输入 OpenAI 兼容 API 基础 URL (例如: https://api.openai.com/v1 或自定义)").strip()
        if not base_url:
            print_error("API 基础 URL 不能为空。")
            return

        # 如果没有协议，添加 https://
        if not base_url.startswith(('http://', 'https://')):
            base_url = "https://" + base_url

        print_info("请输入 API 密钥（输入不会显示在屏幕上）")
        api_key = input_api_key("密钥: ").strip()
        if not api_key:
            print_error("API 密钥不能为空。")
            return

        # 确认输入的密钥（显示部分）
        if not print_api_key_confirmation(hide_api_key(api_key)):
            return

        provider = get_provider("openai", base_url=base_url, api_key=api_key)

    elif provider_id == "iflow":
        # iFlow 配置
        print_info("请输入 iFlow API 密钥（从 https://platform.iflow.cn/profile?tab=apiKey 获取）")
        api_key = input_api_key("密钥: ").strip()
        if not api_key:
            print_error("iFlow API 密钥不能为空。")
            return

        # 确认输入的密钥（显示部分）
        if not print_api_key_confirmation(hide_api_key(api_key)):
            return

        provider = get_provider("iflow", api_key=api_key)

    # 确保 provider 已正确初始化
    if provider is None:
        print_error("AI 供应商初始化失败")
        return

    # 获取可用模型
    available_models = provider.get_models()

    # 选择模型（如果只有一个模型且是Dify，则自动选择）
    if len(available_models) == 1 and "Dify" in provider_name:
        selected_model = available_models[0]
        print(f"自动选择模型: {selected_model}")
    else:
        # 显示所有可用模型
        print_info("可用的模型:")
        for i, model in enumerate(available_models, 1):
            console.print(f"  {i}. {model}", style="white")
        # 添加自定义模型选项
        console.print(f"  {len(available_models) + 1}. 自定义模型", style="cyan")
        console.print()

        while True:
            model_choice = print_input_prompt(f"请选择模型（输入 1-{len(available_models) + 1}）或直接输入模型名")
            try:
                # 尝试将输入作为数字处理
                if model_choice.isdigit():
                    model_num = int(model_choice)
                    if 1 <= model_num <= len(available_models):
                        # 选择预设模型
                        selected_model = available_models[model_num - 1]
                        print_success(f"已选择模型: {selected_model}")
                        break
                    elif model_num == len(available_models) + 1:
                        # 选择自定义模型
                        custom_model = print_input_prompt("请输入自定义模型名称")
                        if custom_model:
                            selected_model = custom_model
                            print_success(f"已选择模型: {selected_model}")
                            break
                        else:
                            print_error("模型名称不能为空，请重新输入。")
                    else:
                        print_error(f"无效的模型序号！请输入 1-{len(available_models) + 1} 之间的数字。")
                else:
                    # 直接输入模型名称（不是数字）
                    if model_choice:
                        selected_model = model_choice
                        print_success(f"已选择模型: {selected_model}")
                        break
                    else:
                        print_error("输入不能为空，请选择模型或输入自定义模型名称。")
            except ValueError:
                print_error("请输入有效的数字或模型名称！")

    # 角色选择
    print_info("可用角色:")
    for i, role in enumerate(ROLES, 1):
        console.print(f"  {i}. {role}", style="bold white")
    console.print(f"  {len(ROLES) + 1}. 自定义角色", style="bold white")
    console.print()

    while True:
        try:
            role_choice = print_input_prompt(f"请选择角色（输入 1-{len(ROLES) + 1}）")

            # 尝试转换为数字
            if role_choice.isdigit():
                role_num = int(role_choice)

                # 选择预设角色
                if 1 <= role_num <= len(ROLES):
                    selected_role = ROLES[role_num - 1]
                    break

                # 自定义角色
                elif role_num == len(ROLES) + 1:
                    while True:
                        custom_role = print_input_prompt("请输入自定义角色名称")
                        if custom_role:
                            selected_role = custom_role
                            break
                        else:
                            print_error("角色名称不能为空，请重新输入。")
                    break

                else:
                    print_error(f"无效的角色序号！请输入 1-{len(ROLES) + 1} 之间的数字。")
            else:
                # 直接输入角色名称
                if role_choice:
                    selected_role = role_choice
                    break
                else:
                    print("输入不能为空，请选择角色或输入自定义角色名称。", file=sys.stderr)

        except ValueError:
            print_error("请输入有效的数字！")
        except KeyboardInterrupt:
            print_warning("用户取消操作，程序退出。")
            sys.exit(0)

    # 模式选择
    # 主循环 - 允许用户多次选择模式
    while True:
        print_info("请选择运行模式:")
        console.print("1. 会话模式 (实时对话)", style="bold white")
        console.print("2. 批量询问模式 (通过 Excel 文件批量询问)", style="bold white")
        console.print("3. 退出程序", style="bold white")
        console.print()
        mode_choice = print_input_prompt("请输入模式序号")

        if mode_choice == '1':
            run_interactive_chat(provider, selected_role, provider_name, selected_model)
        elif mode_choice == '2':
            run_batch_query(provider, selected_role, provider_name, selected_model)
        elif mode_choice == '3':
            print_info("感谢使用，再见！")
            break
        else:
            print_error("无效的模式选择，请重新输入。")
            console.print()
            continue
        
        console.print()
        print_success("已返回主菜单")
        console.print()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n程序已由用户中断。再见！")
